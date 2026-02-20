import streamlit as st
import pandas as pd
import cv2
import numpy as np
from ultralytics import YOLO
import tempfile
import io
import gc
import cv2
import time # [v50 New] Added time import
import os
import traceback
import yaml
import sys
import logging
import uuid # [v15 Fix] Avoid file lock
import random # [v50 New] For diverse AI comments
from datetime import datetime
from collections import defaultdict
import sqlite3 # [v27 Add] Database Support
import mediapipe as mp # [v59 New] MediaPipe Support

# [v27 Add] Database Config
DB_FILE = "hmeayc.db"

def init_db():
    """Initialize SQLite database with required tables."""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # 1. Observations Table (Master)
    c.execute('''CREATE TABLE IF NOT EXISTS observations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    obs_date TEXT,
                    observer_name TEXT,
                    activity_name TEXT,
                    video_file TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_deleted INTEGER DEFAULT 0  -- [v43 New] Soft Delete Flag
                )''')

    # [v43 Fix] Auto-migration check (Add is_deleted if missing)
    try:
        c.execute("SELECT is_deleted FROM observations LIMIT 1")
    except sqlite3.OperationalError:
        c.execute("ALTER TABLE observations ADD COLUMN is_deleted INTEGER DEFAULT 0")
        conn.commit()

    # 2. Records Table (Details per student)
    c.execute('''CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    obs_id INTEGER,
                    student_id TEXT,
                    role TEXT,
                    score REAL,
                    sync_score REAL,
                    focus_score REAL,
                    temp_lag REAL,
                    comment TEXT,
                    FOREIGN KEY (obs_id) REFERENCES observations (id)
                )''')
    conn.commit()
    # [v49 New] Add teacher_grading column if missing (for Decision Tree Prep)
    try:
        c.execute("SELECT teacher_grading FROM records LIMIT 1")
    except sqlite3.OperationalError:
        try:
            c.execute("ALTER TABLE records ADD COLUMN teacher_grading INTEGER DEFAULT 0")
            conn.commit()
        except Exception as e:
            print(f"Migration Error (teacher_grading): {e}")

    # [v58 New] Add stability_score column if missing
    try:
        c.execute("SELECT stability_score FROM records LIMIT 1")
    except sqlite3.OperationalError:
        try:
            c.execute("ALTER TABLE records ADD COLUMN stability_score REAL DEFAULT 0.0")
            conn.commit()
        except Exception as e:
            print(f"Migration Error (stability_score): {e}")

    conn.close()

def save_analysis_to_db(observer, activity, video, df):
    """Save the analyzed dataframe to SQLite with Update Capability."""
    if df.empty: return False

    try:
        conn = sqlite3.connect(DB_FILE) # Fix: Use string or variable consistently
        c = conn.cursor()

        # [v32 Fix] Check if we already have a session ID for this analysis
        obs_id = st.session_state.get('current_obs_id', None)

        if obs_id:
            # Update mode: Update Master Record timestamp
            c.execute("UPDATE observations SET timestamp=CURRENT_TIMESTAMP, is_deleted=0 WHERE id=?", (obs_id,))
            
            # [v21.5 Fix] Instead of deleting ALL, only delete the IDs we are about to re-insert
            # This prevents a single-target report from wiping out the whole group record.
            ids_to_update = df["幼兒 ID"].unique().tolist()
            if ids_to_update:
                placeholders = ', '.join(['?'] * len(ids_to_update))
                query = f"DELETE FROM records WHERE obs_id=? AND student_id IN ({placeholders})"
                c.execute(query, [obs_id] + ids_to_update)
        else:
            # Insert Master Record
            date_str = datetime.now().strftime("%Y-%m-%d")
            c.execute("INSERT INTO observations (obs_date, observer_name, activity_name, video_file, is_deleted) VALUES (?, ?, ?, ?, 0)",
                    (date_str, observer, activity, video))
            obs_id = c.lastrowid
            st.session_state.current_obs_id = obs_id # Save for next time

        # Insert Student Records (New or Re-insert)
        for _, row in df.iterrows():
            # Parse numerical values safe
            try:
                # [v54 Fix] Safe Float Conversion (Handle None/Empty)
                def safe_float(val, default=0.0):
                    try:
                        if val is None or val == "":
                            return default
                        return float(val)
                    except (ValueError, TypeError):
                        return default

                score = safe_float(row.get("AI 觀察判定 (1-5)"))
                sync = safe_float(row.get("跟隨指令 (同步率%)"))
                focus = safe_float(row.get("專注度(%)"))

                # [v58 New] Extract stability score
                stability = safe_float(row.get("動作穩定度"))

                # Parse Lag (string "0.5s" -> float 0.5)
                lag_str = str(row.get("時序延遲 (Lag)", "0")).replace("s", "")
                lag = 0.0
                if lag_str != "-" and lag_str != "None":
                    lag = safe_float(lag_str)

                comment = row.get("AI 總結評語", "")
                role = row.get("參與型態", "Unknown")
                s_id = row.get("幼兒 ID", "Unknown")

                # [v49] Teacher Grading for Decision Tree
                t_grade = safe_float(row.get("教師評分 (1-5)"))

                c.execute('''INSERT INTO records 
                            (obs_id, student_id, role, score, stability_score, sync_score, focus_score, temp_lag, comment, teacher_grading)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (obs_id, s_id, role, score, stability, sync, focus, lag, comment, t_grade))
            except Exception as e:
                print(f"DB Row Error: {e}")
                st.error(f"寫入資料庫失敗 (Row Level): {e}") # [v52 Debug] Show error to user
                pass 

        conn.commit()
        conn.close()
        return True, obs_id # Return ID for feedback
    except Exception as e:
        return False, str(e)

# [v43 New] Soft Delete & Restore Logic
# [v43 New] Soft Delete & Restore Logic
# delete_observation_record relocated below to avoid duplication

def restore_observation_record(obs_id):
    """Restore a soft-deleted record."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("UPDATE observations SET is_deleted=0 WHERE id=?", (obs_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Restore Error: {e}")
        return False
        return False

# [v41 New] History Management Helpers
# [v41 New] History Management Helpers
def delete_student_record(obs_id, student_id):
    """Delete a specific student record from an observation."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("DELETE FROM records WHERE obs_id=? AND student_id=?", (obs_id, student_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Delete Student Error: {e}")
        return False

# [v48 New] Motion Smoothness (Neuro-Motor Analysis)
def calculate_smoothness(pos_history):
    """
    Calculate movement smoothness based on velocity variance (simplified Jerk).
    Returns 0-100 (100 = Very Smooth, 0 = High Jitter/Tremor).
    """
    if len(pos_history) < 5: return 80.0 # Default assumption

    # Calculate velocities (distance between consecutive frames)
    velocities = []
    for i in range(1, len(pos_history)):
        p1, p2 = pos_history[i-1], pos_history[i]
        dist = np.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)
        velocities.append(dist)

    if not velocities: return 80.0

    # Smoothness ≈ Inverse of Velocity Variance (Consistency)
    # If velocity changes abruptly (High Acceleration/Jerk) -> High Variance -> Low Smoothness
    vel_var = np.var(velocities)

    # Normalization (Heuristic)
    # Variance usually 0-1000 depending on movement speed
    # We map 0-500 var to 100-0 score
    smoothness = max(0, 100 - (vel_var / 5.0))
    return round(smoothness, 1)

def rename_student_record(obs_id, old_name, new_name):
    """Rename a specific student in an observation record."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("UPDATE records SET student_id=? WHERE obs_id=? AND student_id=?", (new_name, obs_id, old_name))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Rename Student Error: {e}")
        return False

# [v47 New] Identity Merge for Unifying Records
def merge_student_identity(source_name, target_name):
    """
    Merge all records of 'source_name' into 'target_name'.
    Returns (success, count_updated).
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # Check if source exists
        c.execute("SELECT COUNT(*) FROM records WHERE student_id=?", (source_name,))
        count = c.fetchone()[0]

        if count == 0:
            conn.close()
            return False, 0

        # Update records
        c.execute("UPDATE records SET student_id=? WHERE student_id=?", (target_name, source_name))
        updated_rows = c.rowcount

        conn.commit()
        conn.close()
        return True, updated_rows
    except Exception as e:
        logging.error(f"Merge Identity Error: {e}")
        return False, 0

# [v12] PyInstaller Path Resolver
def get_resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# 設定 logging
logging.basicConfig(filename="app_crash_log.txt", level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s', force=True)

# [v14 New] AI 自動評語生成邏輯
def generate_ai_comment(motion_score, sync_score, actions, gaze_status):
    """
    根據數據生成自然語言評語
    motion_score: 1-5
    sync_score: 0-100 (or None)
    actions: list of strings ['跳躍', '蹲下'...]
    gaze_status: '專注' or '側臉' or '一般'
    """
    comment = ""

    # 1. 活躍度描述
    if motion_score >= 4:
        comment += "表現活力充沛，肢體動作幅度大。"
    elif motion_score == 3:
        comment += "參與度平穩，動作適中。"
    else:
        comment += "處於靜態觀察狀態，動作較少。"

    # 2. 同步率描述 (如果有)
    if sync_score is not None:
        if sync_score >= 80:
            comment += " 與教師動作高度同步，跟隨指令極佳。"
        elif sync_score >= 50:
            comment += " 大致能跟隨教師指令。"
        else:
            comment += " 展現自我風格，未完全跟隨指令。"

    # 3. 動作細節
    if actions:
        unique_actions = list(set(actions))
        action_str = "、".join(unique_actions)
        comment += f" 頻繁出現「{action_str}」等動作。"

    # 4. 專注力
    if gaze_status == "專注":
        comment += " 且全程保持高度專注。"
    elif gaze_status == "側臉":
        comment += " 但注意力似乎較為發散，頻繁轉頭。"

    return comment

try:
    logging.info("App starting...")
except:
    pass
# --- 1. 系統網頁與保險箱設定 ---
st.set_page_config(page_title="HMEAYC 專業觀察系統", layout="wide")
st.title("🩰 解碼教室裡的舞蹈：AI 智能助教系統")
st.caption("最終決賽版：數據同步與視覺強化鎖定模式")
# @st.cache_resource # [重要] 暫時移除快取以清除潛在錯誤狀態
def load_model():
    # 增加錯誤處理
    try:
        # [PyInstaller Fix] Resolve Path
        model_path = get_resource_path("yolov8n-pose.pt")
        model = YOLO(model_path)
        return model
    except Exception as e:
        st.error(f"模型載入失敗: {e}")
        return None
model = load_model()
# 初始化保險箱 (Session State)
if 'id_list' not in st.session_state: st.session_state.id_list = set()
if 'id_features' not in st.session_state: st.session_state.id_features = {}
if 'analysis_done' not in st.session_state: st.session_state.analysis_done = False
if 'last_frame' not in st.session_state: st.session_state.last_frame = None
# 新增：追蹤目前處理完畢的檔案名稱，避免重複跑
if 'processed_file' not in st.session_state: st.session_state.processed_file = None
# [v10] Color Hunt 專業色票庫 (HSV: H[0-180], S[0-255], V[0-255])
REF_COLORS = {
    "正紅": (0, 200, 200),
    "暗紅/棗紅": (175, 160, 100), # [v38 Fix] Capture Dark Red/Maroon
    "酒紅": (170, 200, 100),
    "亮橘": (15, 200, 250),
    "鵝黃": (30, 100, 240),
    "土黃/芥末": (30, 200, 150),
    "米色": (25, 30, 230), # [v38] Lower Hue slightly
    "卡其": (35, 80, 180),
    "深綠": (60, 200, 100),
    "墨綠/軍綠": (55, 90, 60), # [v38 Fix] Lower V (80->60) to avoid confusion with dark red
    "草綠": (50, 200, 200),
    "湖水綠(Teal)": (85, 200, 150),
    "淺藍": (100, 100, 240),
    "牛仔藍": (110, 150, 200),
    "深藍": (115, 200, 100),
    "紫色": (140, 150, 200),
    "淺紫": (135, 90, 220),       
    "粉紅": (160, 100, 240),
    "白/粉紅白": (165, 30, 240), # [v38] Handle pinkish white
    "桃紅/洋紅": (170, 180, 200), 
    "白色": (0, 0, 250), # [v38] Higher V
    "灰色": (0, 0, 128),
    "黑色": (0, 0, 30),
    "焦糖/棕色": (20, 150, 150)
}

def get_dominant_color(img):
    if img.size == 0: return "未知"

    # 轉為 HSV
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)

    # [優化] 去除背景 (簡單用 V > 25 且 S > 10 當作前景)
    valid_mask = (v > 25) 
    if np.count_nonzero(valid_mask) < 10:
        return "黑色" # 幾乎全黑

    avg_h = np.mean(h[valid_mask])
    avg_s = np.mean(s[valid_mask])
    avg_v = np.mean(v[valid_mask])

    current_color = (avg_h, avg_s, avg_v)

    # [演算法] 尋找歐式距離最近的顏色
    min_dist = float('inf')
    best_match = "未知"

    for name, ref_hsv in REF_COLORS.items():
        # H 的距離要特別處理 (因為是環形，0 和 180 很近)
        # 這裡簡化處理：如果 s 很低 (灰色/白色/米色)，H 的權重應該很低

        # 權重調整: V (亮度) 對黑白灰很重要; H (色相) 對彩色很重要; S (飽和) 對米色/卡其中要
        # 這裡使用加權歐式距離
        dh = min(abs(current_color[0] - ref_hsv[0]), 180 - abs(current_color[0] - ref_hsv[0]))
        ds = abs(current_color[1] - ref_hsv[1])
        dv = abs(current_color[2] - ref_hsv[2])

        # 如果是低飽和度 (黑白灰米)，H 的權重要很小
        w_h, w_s, w_v = 1.0, 1.0, 1.0
        if ref_hsv[1] < 50: # 低飽和參考色 (白/灰/米)
            w_h = 0.1 # 色相不重要
            w_s = 2.0 # 飽和度重要 (區分白vs米)
            w_v = 2.0 # 亮度重要 (區分白vs灰vs黑)
        else: # 彩色
            w_h = 2.0 # 色相最重要

        dist = np.sqrt(w_h*(dh**2) + w_s*(ds**2) + w_v*(dv**2))

        if dist < min_dist:
            min_dist = dist
            best_match = name

    # [v8] 深淺前綴 (仍然保留，增加描述性)
    prefix = ""
    # 不對黑白灰加深淺
    if best_match not in ["白色", "黑色", "灰色", "米色"]:
        if avg_v < 80: prefix = "深"
        elif avg_v > 200: prefix = "淺/亮"

    return f"{prefix}{best_match}"


def get_clothing_pattern(img):
    if img.size == 0: return ""
    try:
        # [v9 Enhancement] 進階特徵分析
        # 1. 轉為灰階 & 邊緣檢測
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150)

        # 2. 計算整體紋理密度
        total_pixels = edges.size
        edge_pixels = np.count_nonzero(edges)
        density = edge_pixels / total_pixels

        # 3. 檢查中心區域 (Logo 偵測)
        h, w = gray.shape
        center_h, center_w = int(h*0.3), int(w*0.3)
        center_roi = edges[center_h:h-center_h, center_w:w-center_w]
        center_density = np.count_nonzero(center_roi) / center_roi.size if center_roi.size > 0 else 0

        # 如果中心很複雜 (Density > 0.2) 但整體還好 -> 可能是圖案/Logo
        if center_density > 0.2 and center_density > density * 1.5:
            return "(含圖案/Logo)"

        # 4. 檢查高對比細節
        mask_w = cv2.inRange(gray, 200, 255) # 白色區域
        mask_b = cv2.inRange(gray, 0, 50)    # 黑色區域
        ratio_w = cv2.countNonZero(mask_w) / total_pixels
        ratio_b = cv2.countNonZero(mask_b) / total_pixels

        details = []
        if ratio_w > 0.15: details.append("白") # 超過 15% 是白色
        if ratio_b > 0.15: details.append("黑") # 超過 15% 是黑色

        if details:
            detail_str = "、".join(details)
            if density > 0.15: # 如果同時紋理也複雜
                return f"(含{detail_str}色條紋/花紋)"
            else:
                return f"(含{detail_str}色細節/圖案)"

        # 5. 一般紋理判斷
        if density > 0.2:
            return "(含花紋/條紋)"
        elif density > 0.1:
            return "(含細微花紋)"

        return ""
    except:
        return ""

# ---------------------------
# [v19 New] Advanced Interaction & Focus Helpers
# ---------------------------
def calculate_head_yaw(nose, left_ear, right_ear):
    """
    Estimate head yaw (Left/Right/Center) based on relative ear positions.
    Returns: angle in degrees (approximate), 0=Center, >0=Right, <0=Left
    """
    if nose is None or left_ear is None or right_ear is None:
        return 0 # Unknown

    # Calculate distances
    d_left = np.linalg.norm(np.array(nose) - np.array(left_ear))
    d_right = np.linalg.norm(np.array(nose) - np.array(right_ear))

    total_d = d_left + d_right
    if total_d == 0: return 0

    # Heuristic: If nose is closer to left ear (subject's left), head is turned to subject's left.
    # We want "Observer's View". 
    # Subject Left Ear is typically on Right side of image (if facing front).
    # If Nose moves to Right (closer to Left Ear), Yaw is Positive (Right).

    yaw_factor = (d_right - d_left) / total_d 
    return yaw_factor * 90 # Degrees

def check_gaze_at_target(observer_pos, observer_yaw, target_pos, tolerance=20):
    """
    Check if observer is looking at target (horizontal direction).
    """
    if observer_pos is None or target_pos is None: return False

    dx = target_pos[0] - observer_pos[0]

    # If Target is Right (dx > 0), Observer must look Right (Yaw > threshold)
    # If Target is Left (dx < 0), Observer must look Left (Yaw < -threshold)

    threshold = 10 # Min degrees to be considered "Looking Side"

    if dx > 50: # Target is significantly to the Right
        return observer_yaw > threshold 
    elif dx < -50: # Target is significantly to the Left
        return observer_yaw < -threshold

    # Target is straight ahead (approx same X)
    # Observer should be looking Center
    return abs(observer_yaw) < threshold

def draw_social_graph(interactions, id_map, width=1100, height=1000):
    """
    Draw a social network graph using OpenCV.
    interactions: dict {(id1, id2): count}
    id_map: dict {id: label}
    """
    # Create white canvas
    canvas = np.ones((height, width, 3), dtype=np.uint8) * 255

    # Determine nodes (unique IDs)
    nodes = list(id_map.keys())
    if not nodes: return canvas

    # Position nodes in a circle
    cx, cy = width // 2, height // 2 + 20 # Lower center slightly
    # [v20.11 Layout] Radius 400 (Mega)
    radius = 400 
    node_positions = {}

    for i, node_id in enumerate(nodes):
        angle = 2 * np.pi * i / len(nodes)
        x = int(cx + radius * np.cos(angle))
        y = int(cy + radius * np.sin(angle))
        node_positions[node_id] = (x, y)

    # Draw edges
    max_count = max(interactions.values()) if interactions else 1

    # [v33 Update] Lower threshold to show more connections
    # [v33 Update] Darker lines for visibility

    for (id1, id2), count in interactions.items():
        if count < 2: continue # [v33] Lowered threshold from 3 to 2

        pt1 = node_positions.get(id1)
        pt2 = node_positions.get(id2)

        if pt1 and pt2:
            # Thickness based on frequency (1 to 8)
            thickness = max(1, int((count / max_count) * 8)) 

            # Color: Darker Gray for better contrast (100,100,100)
            cv2.line(canvas, pt1, pt2, (150, 150, 150), thickness)

    # Draw nodes
    for node_id, (x, y) in node_positions.items():
        color = (235, 206, 135) # SkyBlue (BGR)
        # Highlight "Core" nodes (high degree)
        degree = sum([c for (k, c) in interactions.items() if node_id in k])
        if degree > max_count * 0.4: color = (0, 0, 255) # Red (BGR) [v33] Lowered active threshold

        # [v20.12 Refine] Node Radius 30 (User Request: "dots bigger")
        cv2.circle(canvas, (x, y), 30, color, -1) 
        cv2.circle(canvas, (x, y), 30, (0, 0, 0), 2)

        # Show ID
        # [v31] Support Custom Name if available?
        # Canvas drawing logic is separate from DF. 
        # But we can try to look up from session state if simple enough.
        # For graph clarity, maybe keep ID number? Or use Name if short.
        label = str(node_id)
        if "Teacher" in id_map.get(node_id, ""): label = "T"

        # Try to use Custom Name if defined in session state
        # (This is a nice-to-have, but graph might get crowded)
        # Let's check session state
        real_label = label
        # We don't have direct access to session_state here comfortably without import issues?
        # Actually we do, inside the function.
        # But `id_map` passed in is usually just {id: id_string}.

        # [v20.12 Refine] Larger font for larger nodes (0.4 -> 0.8)
        font_scale = 0.8
        ts = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, 2)[0]
        cv2.putText(canvas, label, (x - ts[0]//2, y + ts[1]//2), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), 2)

    return canvas

def get_motion_score(positions):
    if len(positions) < 10: return 3 # 資料不足給平均分
    # 計算座標標準差 (Standard Deviation)
    # 變異數越大表示移動範圍越大
    pos_array = np.array(positions)
    std_x = np.std(pos_array[:, 0])
    std_y = np.std(pos_array[:, 1])
    movement = std_x + std_y

    # 根據移動量給分 (數值需依實際影片調整，這裡先抓個概略值)
    # 根據移動量給分 (數值需依實際影片調整，這裡先抓個概略值)
    if movement > 100: return 5 # 大幅移動
    if movement > 60: return 4 # 明顯移動 (提高門檻)
    if movement > 35: return 3 # 小幅移動 (提高門檻，避免攝影師 3 分)
    if movement > 10: return 2 # 微幅晃動
    return 1 # 幾乎靜止 (可能是背景人物/攝影師)

# [新增] 1. 線性內插 (Interpolation) - 補足短暫消失的軌跡
def interpolate_positions(data_list, max_gap=15):
    # data_list: [(frame_idx, (x, y)), ...]
    if not data_list: return []

    data_list.sort(key=lambda x: x[0]) # 按幀數排序
    interpolated = []

    for i in range(len(data_list) - 1):
        f1, p1 = data_list[i]
        f2, p2 = data_list[i+1]

        interpolated.append((f1, p1))

        gap = f2 - f1
        if 1 < gap <= max_gap:
            # 進行內插
            for j in range(1, gap):
                alpha = j / gap
                new_x = int(p1[0] * (1 - alpha) + p2[0] * alpha)
                new_y = int(p1[1] * (1 - alpha) + p2[1] * alpha)
                new_f = f1 + j
                interpolated.append((new_f, (new_x, new_y)))

    interpolated.append(data_list[-1])
    return interpolated

# [新增] 2. 師生同步率 (Teacher-Student Sync)
# 計算學生與老師的動作向量相似度 (Cosine Similarity)
def calculate_teacher_sync(student_pos, teacher_pos):
    # student_pos, teacher_pos: list of (frame, (x, y))
    # 1. 確保時間對齊 (Intersection of frames)
    s_dict = {f: p for f, p in student_pos}
    t_dict = {f: p for f, p in teacher_pos}

    common_frames = sorted(list(set(s_dict.keys()) & set(t_dict.keys())))
    logging.info(f"Sync Debug: ID={student_pos[0][1]}? Common Frames: {len(common_frames)}")

    if len(common_frames) < 10: 
        logging.info("Sync failed: Not enough common frames (<10)")
        return 0.0 # 重疊時間太短

    # 2. 計算速度向量 (Velocity Vector)
    # v[i] = p[i+1] - p[i]
    s_vecs = []

    for i in range(len(common_frames) - 3): # Skip 3 frames to get better vector
        f1, f2 = common_frames[i], common_frames[i+3]
        if f2 - f1 > 5: continue # 只有連續幀才算向量

        p_s1, p_s2 = s_dict[f1], s_dict[f2]
        p_t1, p_t2 = t_dict[f1], t_dict[f2]

        # 向量 (dx, dy)
        v_s = np.array([p_s2[0] - p_s1[0], p_s2[1] - p_s1[1]])
        v_t = np.array([p_t2[0] - p_t1[0], p_t2[1] - p_t1[1]])

        # 正規化
        norm_s = np.linalg.norm(v_s)
        norm_t = np.linalg.norm(v_t)

        # [v12 Fix] 處理靜止狀態 (Static Handling)
        threshold = 0.5
        is_static_s = norm_s < threshold
        is_static_t = norm_t < threshold

        if is_static_s and is_static_t:
            # 兩者都靜止 -> 視為同步 (給予 1.0 )
            s_vecs.append(1.0)
            continue
        elif is_static_s or is_static_t:
            # 其中一方靜止，另一方在動 -> 視為不同步 (給予 0.0)
            s_vecs.append(0.0)
            continue

        # 兩者都在動 -> 計算向量相似度
        # Cosine Similarity: A . B / |A||B|
        cos_sim = np.dot(v_s, v_t) / (norm_s * norm_t)
        s_vecs.append(cos_sim)

    if not s_vecs: 
        logging.info("Sync failed: No valid frames found")
        return 0.0

    # 3. 平均相似度 (-1 ~ 1) -> 映射到 (0 ~ 100 分)
    avg_sim = np.mean(s_vecs)

    # [Log Debug]
    logging.info(f"Sync Debug: Vectors={len(s_vecs)}, AvgSim={avg_sim:.2f}")

    # 簡單映射：因為我們已經處理了靜止(1.0)和單動(0.0)，剩下的動態部分直接取平均
    # 負值(反向)視為 0
    score = max(0, avg_sim) * 100 

    return round(score, 1)

def analyze_temporal_sync(motion_s, motion_t, fps=30, max_lag_sec=1.5):
    """
    [v26 New] Calculate Temporal Lag using Cross-Correlation on Motion Energy.
    Returns: (max_correlation, lag_in_seconds)
    Msg: "Sync" or "Delay 0.5s"
    """
    if not motion_s or not motion_t: return 0.0, 0.0

    # Ensure equal length / align to the shorter one's end (most recent)
    min_len = min(len(motion_s), len(motion_t))
    if min_len < 30: return 0.0, 0.0 # Too short

    # Take recent history (e.g., last 10 seconds = 300 frames)
    window = 300
    s = np.array(motion_s[-window:])
    t = np.array(motion_t[-window:])

    # Normalize (Z-score) to avoid amplitude bias
    if np.std(s) < 1e-6 or np.std(t) < 1e-6: return 0.0, 0.0
    s = (s - np.mean(s)) / np.std(s)
    t = (t - np.mean(t)) / np.std(t)

    # Cross Correlation
    # mode='full' returns array of size N+M-1. Index of max correlates to lag.
    corr = np.correlate(s, t, mode='full') / len(s) # Normalize by length

    # Lag finding
    # Center index (0 lag) is at len(t) - 1
    lags = np.arange(-len(t) + 1, len(s))
    max_idx = np.argmax(corr)
    max_corr = corr[max_idx]
    best_lag = lags[max_idx]

    # Convert lag frames to seconds
    # lag > 0 means s is AHEAD of t? np.correlate(s, t): sum(s[k] * t[k+delay])
    # Usually: if s is shifted by +d to match t, then s was behind.
    lag_sec = best_lag / fps

    # Limit to realistic lag (e.g. +/- 1.5s)
    if abs(lag_sec) > max_lag_sec:
        return 0.0, 0.0

    return max_corr, lag_sec

# [新增] 2. 同步率 (R-Value) 計算
def calculate_group_sync(id_motion_scores):
    # id_motion_scores: {mid: [score1, score2, ...]}
    # 簡單計算：所有 ID 在同一幀的變異數 (Variance) 的倒數
    # 變異數越小 -> 大家動作越一致 -> 同步率高
    # 這裡簡化為：計算每個人的平均動作分數，然後算這些平均分數的標準差
    # (更精確應該是 Frame-by-frame，但需要對齊時間軸)

    if not id_motion_scores: return 0

    avg_scores = []
    for mid, scores in id_motion_scores.items():
        if scores:
            avg_scores.append(np.mean(scores))

    if len(avg_scores) < 2: return 0 # 只有一人無法算同步

    std_dev = np.std(avg_scores)
    # R值設計：標準差 0 -> R=1; 標準差 2 (大) -> R=0.3
    # 公式：1 / (1 + std_dev)
    r_val = 1 / (1 + std_dev)
    return round(r_val, 2)

def calculate_kuramoto_order_parameter(id_motion_log):
    """
    [v20 New] Compute Group Synchronization using simplified Kuramoto Order Parameter.
    Order Parameter R(t) = |(1/N) * sum(exp(i * theta_j))|
    Here, we approximate phase theta_j using the direction of motion vector.
    """
    # id_motion_log: {mid: [m1, m2, ...]} -> This stores magnitude, not vector.
    # We need vector history. But `id_positions` has (frame, (x,y)).
    # Let's derive velocity vectors from `id_positions`.

def calculate_group_sync(id_positions):
    """
    [v21 Upgrade] Vector Coherence Sync (Directional).
    Computes cosine similarity of motion vectors against the group average vector.
    """
    if not id_positions or len(id_positions) < 2:
        return 0.0

    # 1. Extract Motion Vectors per ID (Frame t vs t-1)
    # We need at least 2 frames of history for each ID
    id_vectors = {} # {id: (mean_vx, mean_vy)}

    all_vectors = []

    for mid, pos_list in id_positions.items():
        if len(pos_list) < 2: continue

        # Calculate recent motion vector (using last few frames)
        # Taking average of last 3 moves to smooth jitter
        recent = pos_list[-5:]
        if len(recent) < 2: continue

        vx_sum, vy_sum = 0, 0
        count = 0
        for i in range(1, len(recent)):
            dx = recent[i][0] - recent[i-1][0]
            dy = recent[i][1] - recent[i-1][1]
            vx_sum += dx
            vy_sum += dy
            count += 1

        if count > 0:
            avg_vx, avg_vy = vx_sum/count, vy_sum/count
            # Normalize to unit vector (Direction only)
            mag = np.sqrt(avg_vx**2 + avg_vy**2)
            if mag > 1.0: # Ignore noise/stationary
                id_vectors[mid] = (avg_vx/mag, avg_vy/mag)
                all_vectors.append((avg_vx/mag, avg_gy/mag))

    if len(all_vectors) < 2: return 0.0

    # 2. Compute Group Average Vector (The "Flow")
    avg_gx = sum(v[0] for v in all_vectors) / len(all_vectors)
    avg_gy = sum(v[1] for v in all_vectors) / len(all_vectors)
    group_mag = np.sqrt(avg_gx**2 + avg_gy**2)

    if group_mag < 0.1: return 0.0 # Group is stationary or canceling out

    # Normalize group vector
    g_unit = (avg_gx/group_mag, avg_gy/group_mag)

    # 3. Compute Coherence (Cosine Similarity)
    coherence_scores = []
    for mid, v in id_vectors.items():
        # Dot product of unit vectors = Cosine Similarity (-1 to 1)
        # We only care about positive sync (0 to 1)
        sim = max(0, v[0]*g_unit[0] + v[1]*g_unit[1])
        coherence_scores.append(sim)

    return round(float(np.mean(coherence_scores)), 2)

# [v30 New] Randomized Advice Templates
ADVICE_TEMPLATES = {
    "active_low_sync": [
        "建議教師邀請其擔任小隊長或示範者，將充沛能量轉化為帶領同儕的動力，提升自信與成就感。",
        "可嘗試賦予其「動作發想者」的角色，在自由律動時間鼓勵其創造新動作，引導同儕模仿。",
        "建議在團體活動中安排「紅綠燈」或「定格」遊戲，幫助其練習肢體控制與衝動抑制。"
    ],
    "passive_observer": [
        "建議可安排與較活躍的同儕配對遊戲，透過同儕互動帶動其肢體開展，從觀察過渡到參與。",
        "教師可主動靠近並以眼神鼓勵，或牽手邀請其加入小圈圈，建立安全感降低焦慮。",
        "建議先從「小道具操作」（如絲巾、響板）入手，降低直接肢體表現的心理門檻。"
    ],
    "delayed_follower": [
        "建議教師在動作示範時提供更明確的視覺提示（如倒數或誇大預備動作），或稍放慢節奏，協助其跟上團體律動。",
        "可運用音樂節拍較明顯的曲目，並搭配口語指令（如「1、2、3、跳」），強化其聽覺與動作的連結。",
        "建議安排其站在教師正對面或視覺干擾較少的位置，確保能清楚接收示範訊號。"
    ],
    "independent": [
        "建議先肯定其獨特的動作表現，再逐步引導其將個人動作融入團體指令，建立連結感。",
        "在團體指令中保留「自由發揮」的時段，滿足其探索需求，同時要求在特定訊號回到團體規範。",
        "可邀請其在特定的曲目中擔任「小老師」帶領大家做自己的動作，增進參與感。"
    ],
    "low_focus": [
        "建議運用色彩鮮豔的教具或聲音變化來吸引其視覺注意力，增加眼神接觸的機會。",
        "教師可縮短指令語句，並在發號施令前先喊其名字或輕拍肩膀，確保注意力已連線。",
        "建議安排在前排或靠近教師的位置，減少環境視覺干擾，並給予即時的正向回饋。"
    ],
    "high_performance": [
        "建議提供更高難度或變化的動作挑戰，維持其學習動機並展現優勢能力。",
        "可邀請其擔任「小老師」協助其他同學，培養同理心與照顧他人的能力。",
        "建議引導其注意動作的細節與質感（如「輕輕地飛」vs「用力地跳」），深化美感體驗。"
    ]
}

def generate_expert_comment(score, sync_score, focus_score, role, valid_tags, class_stats=None, archetype_text=""):
    """
    [v21 Upgrade] Context-Aware Expert System with Non-Arbitrary Language.
    [v26 Update] Added Temporal Sync Logic.
    [v30 Update] Randomized Advice for Variety.
    [v61 Update] Merging Archetype Text into the final comment block.
    """
    parts = []

    # [v26] Extract Temporal Info
    temp_corr = 0.0
    lag_sec = 0.0
    if class_stats:
        temp_corr = class_stats.get('temp_corr', 0.0)
        lag_sec = class_stats.get('lag_sec', 0.0)

    # [Fix] Sanitize inputs
    if sync_score is None: sync_score = 0
    if focus_score is None: focus_score = 0

    # 1. Role-based Observation (Descriptive)
    if not role or role == "Unknown":
        pass
    elif "Teacher" in role:
        parts.append("擔任示範者角色，動作引導清晰。")
    elif "Active" in role:
        parts.append("展現高度動作活躍性，積極參與活動。")
    elif "Passive" in role:
        parts.append("維持高互動頻率，但肢體動作幅度較小（靜態互動）。")
    elif "Focused" in role:
        parts.append("動作幅度適中，視覺注意力高度集中於示範者。")
    elif "Imitating" in role:
        parts.append("展現明顯的模仿行為，與示範者動作同步度高。")
    elif "Independent" in role:
        parts.append("表現出獨立的動作模式，較少跟隨示範者。")

    # 2. Sync / Temporal Analysis (Descriptive)
    if sync_score > 80:
        parts.append("動作與示範者高度同步，節奏掌握精確。")
    elif sync_score < 60 and temp_corr > 0.6:
        # Delayed Follower
        if abs(lag_sec) > 0.2:
            parts.append(f"觀察到動作有約 {abs(lag_sec):.2f} 秒的延遲，顯示其為觀察後模仿的習性。")
        else:
            parts.append("雖空間同步率較低，但時間序列顯示其動作趨勢與老師一致。")
    elif sync_score < 40:
        parts.append("動作節奏與團體有顯著差異，可能有自己的探索步調。")

    # 4. Action Specifics
    # Filter out internal tags
    visible_actions = [t for t in valid_tags if t not in ['專注', '側臉']]
    if visible_actions:
        action_str = "、".join(visible_actions)
        parts.append(f"頻繁出現「{action_str}」等動作特徵。")

    if archetype_text:
        parts.append(archetype_text)

    # 5. Teacher Guidance (New)
    # Provide actionable advice based on profile
    suggestions = []
    import random # Ensure random is available

    # Case: Active but Low Sync (High Energy, Independent/Active)
    if "Active" in role and sync_score < 60:
        suggestions.append(random.choice(ADVICE_TEMPLATES["active_low_sync"]))

    # Case: Passive but High Focus (Observer)
    elif "Passive" in role or ("Focused" in role and score < 3):
        suggestions.append(random.choice(ADVICE_TEMPLATES["passive_observer"]))

    # Case: Delayed Sync (Slow Follower)
    elif abs(lag_sec) > 0.3:
        suggestions.append(random.choice(ADVICE_TEMPLATES["delayed_follower"]))

    # Case: Independent (Low Sync, Own Pace)
    elif "Independent" in role and sync_score < 40:
        suggestions.append(random.choice(ADVICE_TEMPLATES["independent"]))

    # Case: Low Focus (Distracted)
    elif focus_score < 40:
        suggestions.append(random.choice(ADVICE_TEMPLATES["low_focus"]))

    # Case: High Performance (High Sync + High Focus)
    elif sync_score > 80 and focus_score > 80:
        suggestions.append(random.choice(ADVICE_TEMPLATES["high_performance"]))

    if suggestions:
        parts.append("\n\n💡 教學建議：" + "".join(suggestions))

    return "".join(parts)

# ... (Original detectaction_and_gaze stays above, skipping diff context here for brevity) ...

# ... (Previous code) ...

# [v30 New] History Management
def delete_observation_record(obs_id):
    """Soft Delete: Mark as deleted but keep data."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        # Soft delete only master record is enough if we filter by it
        c.execute("UPDATE observations SET is_deleted=1 WHERE id=?", (obs_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Soft Delete Error: {e}")
        return False

def show_history_ui():
    st.title("🗄️ 歷史紀錄與幼兒成長歷程")

    # 1. Fetch Master List (Filter OUT deleted items)
    conn = sqlite3.connect(DB_FILE)
    df_obs = pd.read_sql_query("SELECT * FROM observations WHERE is_deleted=0 ORDER BY timestamp DESC", conn)
    conn.close()

    if df_obs.empty:
        st.info("尚無歷史紀錄 (或全部已刪除)。請先在「全功能分析」模式下進行分析並儲存。")
        # Allow viewing Trash Can even if main list is empty?
        # Maybe not block immediately if we want to access Trash Can.
        # But for now, user might have deleted everything.
        # Let's check if there are ANY records including deleted.
        conn = sqlite3.connect(DB_FILE)
        count = pd.read_sql_query("SELECT COUNT(*) FROM observations", conn).iloc[0,0]
        conn.close()

        if count == 0:
            st.stop()
        else:
            st.warning("目前顯示列表為空，但垃圾桶中可能有資料。")

    # Master Table
    st.subheader("📋 活動觀察紀錄列表")

    # [v41] Manage/Delete Section (Enhanced)
    with st.expander("🛠️ 資料庫管理 (Database Management)"):
        # Create tabs for different management actions
        m_tab1, m_tab2, m_tab3, m_tab4 = st.tabs(["🗑️ 刪除整筆觀察", "✏️ 修改/刪除幼兒資料", "♻️ 垃圾桶 (復原刪除)", "👥 身份合併 (Merge)"])

        with m_tab1:
            st.caption("⚠️ 此操作將刪除該次觀察的所有數據 (包含所有幼兒紀錄)。")
            obs_to_delete = st.selectbox("選擇要刪除的紀錄:", 
                                    df_obs['id'].astype(str) + " | " + df_obs['obs_date'] + " | " + df_obs['activity_name'],
                                    index=None,
                                    placeholder="請選擇...",
                                    key="del_obs_select"
            )
            if obs_to_delete:
                obs_id = int(obs_to_delete.split(" | ")[0])
                if st.button(f"確認刪除紀錄 ({obs_id})", type="primary", key="btn_del_obs"):
                    if delete_observation_record(obs_id):
                        st.success("紀錄已刪除！請重新整理頁面。")
                        st.session_state.clear()
                        st.rerun()
                    else:
                        st.error("刪除失敗。")

        with m_tab2:
            st.caption("🔧 針對特定幼兒進行改名或刪除操作。")

            # 1. Select Observation
            target_obs_str = st.selectbox("1. 選擇紀錄:", 
                                    df_obs['id'].astype(str) + " | " + df_obs['obs_date'] + " | " + df_obs['activity_name'],
                                    index=None,
                                    placeholder="請先選擇紀錄...",
                                    key="edit_obs_select"
            )

            if target_obs_str:
                target_obs_id = int(target_obs_str.split(" | ")[0])

                # 2. Get Students in this observation
                conn = sqlite3.connect("hmeayc.db")
                stu_df = pd.read_sql_query("SELECT DISTINCT student_id FROM records WHERE obs_id=?", conn, params=(target_obs_id,))
                conn.close()

                target_student = st.selectbox("2. 選擇幼兒:", 
                                            stu_df['student_id'].tolist(),
                                            index=None,
                                            placeholder="請選擇幼兒...",
                                            key="edit_stu_select"
                )

                if target_student:
                    action = st.radio("3. 選擇操作:", ["改名 (Rename)", "刪除此人 (Delete)"], horizontal=True)

                    if action == "改名 (Rename)":
                        new_name_input = st.text_input("輸入新名稱:", value=target_student)
                        if st.button("確認改名", key="btn_rename"):
                            if new_name_input and new_name_input != target_student:
                                if rename_student_record(target_obs_id, target_student, new_name_input):
                                    st.success(f"已將 {target_student} 改名為 {new_name_input}")
                                    st.rerun()
                                else:
                                    st.error("改名失敗")
                            else:
                                st.warning("名稱未變更")

                    elif action == "刪除此人 (Delete)":
                        st.error(f"⚠️ 即將從此紀錄中移除 {target_student}，此操作僅影響單一學生！")
                        if st.button("確認移除此人", type="primary", key="btn_del_stu"):
                            if delete_student_record(target_obs_id, target_student):
                                st.success(f"已移除 {target_student}")
                                st.rerun()
                            else:
                                st.error("移除失敗")


        with m_tab3: # [v43 New] Trash Can Tab
            st.caption("♻️ 這裡存放被軟刪除的紀錄，您可以隨時復原。")
            conn = sqlite3.connect(DB_FILE)
            df_deleted = pd.read_sql_query("SELECT * FROM observations WHERE is_deleted=1 ORDER BY timestamp DESC", conn)
            conn.close()

            if df_deleted.empty:
                st.info("垃圾桶是空的 (0 筆資料)。")
            else:
                obs_to_restore = st.selectbox("選擇要復原的紀錄:", 
                                        df_deleted['id'].astype(str) + " | " + df_deleted['obs_date'] + " | " + df_deleted['activity_name'],
                                        index=None,
                                        placeholder="請選擇復原對象...",
                                        key="restore_obs_select"
                )
                if obs_to_restore:
                    obs_id_restore = int(obs_to_restore.split(" | ")[0])
                    if st.button(f"確認復原紀錄 ({obs_id_restore})", key="btn_restore"):
                        if restore_observation_record(obs_id_restore):
                            st.success(f"紀錄 {obs_id_restore} 已成功復原！")
                            st.rerun()
                        else:
                            st.error("復原失敗")

        with m_tab4: # [v47 New] Identity Merge Tab
            st.caption("👥 將多個暫存 ID (如 ID_1) 合併到同一位學生 (如 小明) 名下。此操作無法復原。")

            # Get distinct student IDs
            conn = sqlite3.connect(DB_FILE)
            temp_df = pd.read_sql_query("SELECT DISTINCT student_id FROM records ORDER BY student_id", conn)
            conn.close()

            all_ids = temp_df['student_id'].tolist() if not temp_df.empty else []

            col_m1, col_m2 = st.columns(2)

            with col_m1:
                target_merge_name = st.selectbox("1. 保留的目標 (Target):", all_ids, key="merge_target", index=0 if all_ids else None)

            with col_m2:
                # Filter out target from source options
                source_options = [x for x in all_ids if x != target_merge_name]
                source_merge_name = st.selectbox("2. 要合併的來源 (Source):", source_options, key="merge_source", index=0 if source_options else None)

            if st.button("🚀 確認合併身分", type="primary", use_container_width=True, disabled=not (target_merge_name and source_merge_name)):
                if target_merge_name == source_merge_name:
                    st.warning("目標與來源不能相同。")
                else:
                    success, count = merge_student_identity(source_merge_name, target_merge_name)
                    if success:
                        st.success(f"✅ 成功將 {count} 筆紀錄從 '{source_merge_name}' 合併至 '{target_merge_name}'！")
                        st.rerun()
                    else:
                        st.error("❌ 合併失敗或來源無資料。")

    st.dataframe(df_obs, use_container_width=True, hide_index=True)

    # ... (Rest of existing history UI logic: Student Selection, Charts etc.) ...
    # Re-implementing the rest of show_history_ui below to ensure continuity

    st.markdown("---")
    st.subheader("📈 幼兒個人成長歷程")

    # Get unique student IDs from all records
    conn = sqlite3.connect(DB_FILE)
    all_students = pd.read_sql_query("SELECT DISTINCT student_id FROM records ORDER BY student_id", conn)
    conn.close()

    student_list = all_students['student_id'].tolist()

    # Helper to sort IDs numerically if possible
    def try_sort(x):
        try:
            if "ID_" in x and "(" in x: # Format: ID_1 (Original)
                return int(x.split("_")[1].split(" ")[0])
            return x
        except:
            return x

    # student_list.sort(key=try_sort) # Sort tricky with mixed types

    selected_student = st.selectbox("選擇幼兒 (ID/姓名):", student_list)

    if selected_student:
        # Fetch history for this student
        # JOIN to get date/activity
        q = f"""
        SELECT r.*, o.obs_date, o.activity_name 
        FROM records r
        JOIN observations o ON r.obs_id = o.id
        WHERE r.student_id = '{selected_student}'
        ORDER BY o.obs_date ASC
        """
        conn = sqlite3.connect(DB_FILE)
        df_hist = pd.read_sql_query(q, conn)
        conn.close()

        if not df_hist.empty:
            # Metrics
            col1, col2 = st.columns(2)
            avg_sync = df_hist['sync_score'].mean()
            avg_focus = df_hist['focus_score'].mean()

            col1.metric("平均同步率", f"{avg_sync:.1f}%")
            col2.metric("平均專注度", f"{avg_focus:.1f}%")

            # Charts
            tab1, tab2 = st.tabs(["📊 趨勢圖表", "📝 詳細數據"])

            with tab1:
                # 1. Sync & Focus over time
                chart_data = df_hist[['obs_date', 'sync_score', 'focus_score']].set_index('obs_date')
                st.line_chart(chart_data)

                # 2. Activity / Motion Score
                st.caption("動作活躍度 (1-5) 變化")
                st.bar_chart(df_hist[['obs_date', 'score']].set_index('obs_date'))

            with tab2:
                # [v39 Fix] Rename columns for display
                # [v58] Added stability_score
                display_df = df_hist[['obs_date', 'activity_name', 'role', 'score', 'stability_score', 'sync_score', 'focus_score', 'comment']].copy()
                
                # [v21.4 Refine] Round to 1 decimal place as requested
                for col in ['score', 'stability_score', 'sync_score', 'focus_score']:
                    display_df[col] = display_df[col].round(1)

                display_df.columns = ["日期", "活動名稱", "參與型態", "AI 評分(活躍)", "動作穩定度", "同步率", "專注度", "AI 總結評語"]
                display_df.index = display_df.index + 1 # [v55 Fix] Start index from 1 for user friendly display
                st.table(display_df) # [v47 Fix] Use st.table to allow text wrapping for long comments
        else:
            st.warning("此幼兒尚無詳細數據。")

    # Stop execution here to prevent Main UI from rendering below
    st.stop()

# [v13 New] 動作與視線規則檢測 (增加 跳躍/躺下)
def detectaction_and_gaze(kpts, bbox=None): # 新增 bbox 參數用於長寬比判斷
    """
    kpts: (17, 3) array [x, y, conf]
    bbox: [x1, y1, x2, y2]
    Keypoints:
    0: Nose, 1: LEye, 2: REye, 3: LEar, 4: REar
    5: LSho, 6: RSho, 7: LElb, 8: RElb, 9: LWri, 10: RWri
    11: LHip, 12: RHip, 13: LKne, 14: RKne, 15: LAnk, 16: RAnk
    """
    actions = []

    # 1. 舉手 (Hands Up): 手腕 (9/10) 高於 眼睛 (1/2) 或 耳朵 (3/4)
    # 注意: Y 軸向下為正，所以 "高於" 是 y < target_y
    # 先做信心過濾
    if kpts[9, 2] > 0.5 and kpts[1, 2] > 0.5:
        if kpts[9, 1] < kpts[1, 1]: actions.append("舉手")
    elif kpts[10, 2] > 0.5 and kpts[2, 2] > 0.5:
        if kpts[10, 1] < kpts[2, 1]: actions.append("舉手")

    # 2. 蹲下 (Squat): 臀部 (11/12) 與 膝蓋 (13/14) 垂直距離縮短
    # ... (原有蹲下邏輯) ...
    # 2. 蹲下 (Squat): 臀部 (11/12) 與 膝蓋 (13/14) 垂直距離縮短
    # [v29 Fix] Lower confidence threshold (0.5 -> 0.3) for crowded scenes where legs are occluded
    if kpts[11, 2] > 0.3 and kpts[15, 2] > 0.3 and kpts[5, 2] > 0.3:
        leg_len = kpts[15, 1] - kpts[11, 1]
        body_len = kpts[15, 1] - kpts[5, 1] # 肩到腳
        # [v29 Fix] Relax ratio (0.35 -> 0.45) to detect squats even from high angles
        if body_len > 0 and (leg_len / body_len) < 0.45: 
            actions.append("蹲下")

    # [v13] 3. 躺下/地板動作 (Lying/Floor)
    if bbox is not None:
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        if w > h * 1.2: # 寬大於高 1.2 倍
            actions.append("地板動作")

    # [v13] 4. 跳躍/抬腿 (Jump/High Knees)
    # 邏輯：雙腳腳踝 (15/16) 的 Y 座標小於 (高於) 膝蓋 (13/14) 
    # 或者 腳踝非常接近膝蓋水平
    # 簡單版：雙腳騰空 (Ankle < Knee + offset)
    if kpts[15, 2] > 0.3 and kpts[16, 2] > 0.3 and kpts[13, 2] > 0.3 and kpts[14, 2] > 0.3:
        # 檢查左腳
        l_high = kpts[15, 1] < (kpts[13, 1] + 20) # 腳踝高於膝蓋附近
        r_high = kpts[16, 1] < (kpts[14, 1] + 20)

        if l_high and r_high:
            actions.append("跳躍") # 雙腳都高
        elif l_high or r_high:
            actions.append("抬腿") # 單腳

    # 5. 視線 (Gaze): 耳朵對稱性
    # ... (原有視線邏輯) ...
    if kpts[3, 2] > 0.3 and kpts[4, 2] > 0.3:
        # 兩耳都看得到 -> 正臉/專注
        actions.append("專注")
    elif (kpts[3, 2] > 0.5 and kpts[4, 2] < 0.1) or (kpts[3, 2] < 0.1 and kpts[4, 2] > 0.5):
        # 只有一邊耳朵 -> 側臉
        actions.append("側臉")

    return list(set(actions)) # 去重

# [New Helper UI]
def create_tracker_config():
    config_content = """
tracker_type: botsort
track_high_thresh: 0.25  # [最終修正] 0.25 (捕捉跳動/低品質偵測，挽救 ID 16/18/20)
track_low_thresh: 0.1
new_track_thresh: 0.45   # [救回 ID 18] 降回 0.45，讓短暫出現的人能被追蹤
track_buffer: 60  
match_thresh: 0.7        # [最終修正] 維持 0.7 平衡
fuse_score: True
gmc_method: sparseOptFlow
proximity_thresh: 0.5
appearance_thresh: 0.25
with_reid: False
"""
    try:
        with open("custom_botsort_v6.yaml", "w", encoding="utf-8") as f: # v6
            f.write(config_content)
        return "custom_botsort_v6.yaml"
    except:
        return "botsort.yaml" # Fallback

def get_color_histogram(img):
    if img.size == 0: return None
    # 計算 HSV 直方圖作為 Re-ID 特徵
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    hist = cv2.calcHist([hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])
    cv2.normalize(hist, hist, 0, 1, cv2.NORM_MINMAX)
    return hist

# [v21.5.2] Stability Release
st.sidebar.caption("Version: v21.5.2-FullRes")

# [v65 New] Callback to reset analysis state when tracking settings change
def reset_analysis_state():
    """Callback to unlock the 'Start Analysis' button and clear report when settings change."""
    st.session_state.processed_file = None
    st.session_state.analysis_done = False
    
    # [v21.4.3 Fix] Truly delete result keys to avoid st.download_button NoneType error
    keys_to_clear = [
        'id_list', 'id_interactions', 'id_positions', 
        'final_id_count', 'final_id_list', 'excel_ready_data', 
        'group_sync_r', 'social_graph_image', 'video_output_path'
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
            
    # Re-initialize defaults if needed by other components
    st.session_state.id_list = set()
    st.session_state.id_interactions = defaultdict(int) 
    st.session_state.id_positions = {}

# --- 2. 側邊欄：資訊固定 ---
st.sidebar.header("📋 基本資訊輸入")
observer_name = st.sidebar.text_input("觀察員姓名", "文禎")
act_name = st.sidebar.text_input("活動名稱", "Walk and copy animal")
act_date = st.sidebar.date_input("觀察日期", datetime.now())
music_element = st.sidebar.text_input("音樂元素 (如：走停、快慢)", "走停")

# [v47 New] Performance Mode
perf_mode = st.sidebar.selectbox(
    "⚗️ 分析效能模式", 
    ["⚡ 標準模式 (Balanced)", "🚀 疾速模式 (Turbo)", "🎯 精準模式 (Pro)", "🔬 微觀分析模式 (MediaPipe Holistic)"],
    index=0,
    help="選擇分析頻率以平衡速度與精準度。微觀模式專用於分析單一個人。",
    on_change=reset_analysis_state
)

if "Turbo" in perf_mode:
    frame_interval = 4
    st.sidebar.caption("🚀 每 4 幀取樣 1 次 (極速，適合快速瀏覽)")
elif "Pro" in perf_mode:
    frame_interval = 1
    st.sidebar.caption("🎯 每 1 幀都分析 (最慢，捕捉最細微動作)")
elif "MediaPipe" in perf_mode:
    frame_interval = 1
    st.sidebar.caption("🔬 啟用 MediaPipe 擷取單一個人 540+ 關鍵點")
else:
    frame_interval = 2
    st.sidebar.caption("⚡ 每 2 幀取樣 1 次 (預設，平衡速度與準確)")

st.sidebar.markdown("---")
st.sidebar.markdown("### 🎯 追蹤目標設定 (全模式通用)")

# [v65 Fix] Dynamic Max ID based on previous run results
max_id = 100  # Default before any run
if 'final_id_list' in st.session_state and st.session_state.final_id_list:
    max_id = max(st.session_state.final_id_list)

# Safely bound the initial value
default_val = 1 if "MediaPipe" in perf_mode else 0
if default_val > max_id:
    default_val = max_id

# [v21.4.4] Persistence Fix: Use a key to prevent the value from resetting to 0 during reruns
target_track_id = st.sidebar.number_input(
    "強制鎖定特定幼兒 ID", 
    min_value=0, max_value=max(1, max_id), 
    step=1, 
    key="locked_target_id",
    help="輸入您想觀察的幼兒ID。若設為 0，則全功能模式會追蹤所有人。",
    on_change=reset_analysis_state
)
# [v48 New] Context-Aware Sync (Phase 5)
st.sidebar.markdown("### 🎭 活動性質設定")

# [v52 Fix] Move Mode Selection to Top to prevent State Loss on Re-run
mode = st.sidebar.radio("模式選擇", ["🚀 全功能分析", "🗄️ 歷史紀錄查閱"], index=0, key="nav_mode")

activity_context = st.sidebar.radio(
    "選擇活動類型:",
    ["跟隨模仿 (Imitation)", "自由創作 (Creative)"],
    index=0,
    help="「自由創作」模式下，系統會包容低同步率，不將其視為異常。"
)
st.session_state.activity_context = activity_context

# [v51 New] Decision Tree Visualization Button
if st.sidebar.button("🌳 顯示 AI 決策樹邏輯"):
    st.session_state.show_decision_tree = True

# [v51 New] Show Decision Tree Diagram (Safe Version)
if st.session_state.get("show_decision_tree", False):
    st.info("🌳 這是目前系統採用的「專家規則決策樹 (Expert Rule Decision Tree)」，模擬了老師的判斷邏輯。")
    # ... (Markdown content is same) ...
    st.markdown("""
    #### 🌲 專家邏輯可視化 (Decision Logic)
    ```mermaid
    graph LR
        A[開始分析] --> B{活動類型?}
        B -->|Creative| C[自由創作]
        B -->|Imitation| D[跟隨模仿]

        C -->|Sync < 40%| H[✨ 展現獨創性]
        C -->|Sync 40-70%| I[🎨 部分跟隨]
        C -->|Sync > 70%| J[🤝 仍舊跟隨]

        D -->|Sync > 80%| O[🌟 高度同步]
        D -->|Sync > 60%| P[✅ 良好跟隨]
        D -->|Sync > 40%| Q[⚠️ 部分分心]
        D -->|Sync < 40%| R[❌ 脫離群體]
    ```
    *(若圖表未顯示，代表您的瀏覽器暫未支援 Mermaid 渲染，但不影響系統運作)*
    """)

    if st.button("關閉決策樹說明"):
        st.session_state.show_decision_tree = False
        st.rerun()

# [v27] Init DB
try:
    init_db()
except Exception as e:
    st.sidebar.error(f"DB Init Error: {e}")

if mode == "🗄️ 歷史紀錄查閱":
    show_history_ui()
    st.stop() # Stop execution to hide analysis UI



if mode == "🗄️ 歷史紀錄查閱":
    show_history_ui()
    st.stop() # Stop execution to hide analysis UI

# --- 3. 影片分析區 ---
# [v65 Critical Fix] Move file_uploader OUTSIDE the conditional block.
# If a Streamlit widget is hidden during a rerun, its state (the uploaded file) is destroyed.
uploaded_file = st.file_uploader("📤 上傳影片 (分析時 ID 將自動歸 1)", type=["mp4", "mov"])

if uploaded_file:
    # 檢查是否為新檔案，如果是則重置
    if 'current_fn' not in st.session_state or st.session_state.current_fn != uploaded_file.name:
        st.session_state.current_fn = uploaded_file.name # [修正] 必須更新 current_fn，否則無限重置
        st.session_state.id_list = set()
        st.session_state.id_features = {}
        st.session_state.id_tracking_count = {} 
        st.session_state.id_positions = {} # 新增：記錄每個 ID 的位置歷程 [(frame_idx, (x, y)), ...]
        st.session_state.id_motion_log = {} # 新增：記錄每個 ID 的動作分數歷程 {mid: [score, ...]}
        st.session_state.id_actions = defaultdict(lambda: defaultdict(int)) # [v12] Action Tracking
        st.session_state.processed_file = None 
        st.session_state.last_frame = None
        st.session_state.lost_ids = {} # {id: {'hist': hist, 'last_seen': frame_idx, 'feat': clothing_str}}
        st.session_state.id_map = {}   # {temp_id: real_id} mapping
        st.session_state.final_id_count = 0
        st.session_state.reindexing_done = False # [v8 Fix] 防止重複重新編號
        st.session_state.final_id_list = []

        # [v19 New] Advanced Analytics State
        st.session_state.id_yaw_history = {} # {mid: [yaw1, yaw2...]}
        st.session_state.id_focus_score = {} # {mid: focus_frames}
        st.session_state.id_interactions = defaultdict(int) # {(id1, id2): count}
        st.session_state.id_gaze_start = {} # [v19.1]
        st.session_state.social_graph_image = None

        # [v48 New] Reset Smoothness Log
        st.session_state.id_smoothness_log = defaultdict(list)

        if model and hasattr(model, 'predictor') and model.predictor is not None:
            # model.predictor.trackers = [] 
            pass
            
        # [v21.4 Fix] Persist Video immediately upon upload to prevent NoneType error later
        try:
            temp_dir = tempfile.gettempdir()
            # Clean up old persistent files to save space
            for f in os.listdir(temp_dir):
                if f.startswith("hmeayc_persist_"):
                    try: os.remove(os.path.join(temp_dir, f))
                    except: pass
                    
            p_path = os.path.join(temp_dir, f"hmeayc_persist_{uuid.uuid4().hex[:8]}.mp4")
            uploaded_file.seek(0)
            with open(p_path, "wb") as f:
                f.write(uploaded_file.read())
            st.session_state.current_tfile_path = p_path
            st.session_state.current_fn = uploaded_file.name
        except Exception as e:
            st.error(f"⚠️ 影片預處理失敗: {e}")

        # 清除錯誤位置的定義
        gc.collect()

if not st.session_state.analysis_done:
    # 只有當「尚未處理過」這個檔案時，才執行分析
    if st.session_state.get('current_fn') and st.session_state.processed_file != st.session_state.current_fn:

        # 顯示開始按鈕，讓使用者準備好再跑
        if st.button("🚀 開始 AI 辨識分析"):
            if model is None:
                st.error("❌ 模型未正確載入，無法執行分析。")
                st.stop()
            # [v21.4 Fix] Use the persisted file path primarily to avoid NoneType/Buffer issues
            tfile_path = st.session_state.get('current_tfile_path')
            if not tfile_path or not os.path.exists(tfile_path):
                # Final fallback: if somehow path is missing but file is there
                if uploaded_file is not None:
                    tfile_path = os.path.join(tempfile.gettempdir(), f"hmeayc_final_{uuid.uuid4().hex[:8]}.mp4")
                    uploaded_file.seek(0)
                    with open(tfile_path, "wb") as f:
                        f.write(uploaded_file.read())
                    st.session_state.current_tfile_path = tfile_path
                else:
                    st.error("❌ 找不到原始影片檔案 (Buffer Lost)，請點擊 [🗑️ 清空] 後重新上傳。")
                    st.stop()

            st.info(f"Debug Info: 暫存檔路徑 = {tfile_path}")

            # [v16 Fix] 強制重置追蹤狀態，確保多次執行不會累積舊資料
            st.session_state.id_list = set()
            st.session_state.id_features = {}
            st.session_state.id_tracking_count = {} 
            st.session_state.id_positions = {}
            st.session_state.id_motion_log = {}
            st.session_state.id_actions = defaultdict(lambda: defaultdict(int))
            st.session_state.lost_ids = {}
            st.session_state.id_map = {}
            st.session_state.display_mapping = {}
            st.session_state.final_id_count = 0
            st.session_state.final_id_count = 0
            st.session_state.video_output_path = None # Reset previous video path

            # [v19 New] Reset Advanced State
            st.session_state.id_yaw_history = {}
            st.session_state.id_focus_score = {}
            st.session_state.id_interactions = defaultdict(int)
            st.session_state.id_gaze_start = {} # [v19.1] {(id1, id2): start_frame}
            st.session_state.social_graph_image = None

            # [v48 New] Reset Smoothness Log
            st.session_state.id_smoothness_log = defaultdict(list)

            try:
                cap = cv2.VideoCapture(tfile_path)

                if not cap.isOpened():
                    st.error(f"❌ 無法開啟影片檔案: {tfile_path}")
                else:
                    # [v15 New] Video Writer for Replay
                    # [v47 Fix] Dynamic FPS based on Performance Mode
                    fps = int(cap.get(cv2.CAP_PROP_FPS)) // frame_interval 
                    if fps < 1: fps = 1
                    st.session_state.effective_fps = fps # [v47] Persist for Post-Analysis
                    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

                    # [v15 Fix] Use UUID to prevent file locking
                    unique_id = uuid.uuid4().hex[:8]
                    # [v18.6 Fix] Save to project dir to ensure persistence
                    output_path = os.path.abspath("obs_video.mp4")

                    # [v23 Fix] Revert to 'avc1' first for Local Compatibility
                    # If 'avc1' fails (e.g. Cloud), it falls back to 'mp4v'
                    # Then FFmpeg Post-Processing (below) fixes it for Web
                    fourcc = cv2.VideoWriter_fourcc(*'avc1') 
                    try:
                        out_video = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
                        if not out_video.isOpened():
                            # Fallback to mp4v if avc1 fails
                            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                            out_video = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
                    except Exception as e:
                        st.warning(f"⚠️ 影片寫入器初始化失敗: {e}")
                        out_video = None

                    # st.write("Debug Info: 影片開啟成功")

                # [v18 Fix] Restore st_frame which was accidentally removed
                st_frame = st.empty()
                st_progress = st.progress(0)
                log_container = st.empty() # 用於顯示即時 debug 訊息

                # [v16] Initialize Display Mapping
                st.session_state.display_mapping = {} 

                st.info("AI 辨識中... 本次分析 ID 將從 1 開始編號 (已啟用 ID 重映射)。")

                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                st.write(f"Debug Info: 總幀數 = {total_frames}")

                if total_frames <= 0: total_frames = 1000 # 防呆
                f_idx = 0

                # [v18.1 Fix] Ensure st_frame is defined in correct scope
                st_frame = st.empty()

                # [v59] Setup MediaPipe Holistic Model
                use_mp = "MediaPipe" in perf_mode
                mp_holistic = None
                mp_drawing = None
                mp_drawing_styles = None
                holistic_model = None
                if use_mp:
                    mp_holistic = mp.solutions.holistic
                    mp_drawing = mp.solutions.drawing_utils
                    mp_drawing_styles = mp.solutions.drawing_styles
                    try:
                        holistic_model = mp_holistic.Holistic(
                            static_image_mode=False,
                            model_complexity=1, 
                            enable_segmentation=False,
                            refine_face_landmarks=True, 
                            min_detection_confidence=0.5,
                            min_tracking_confidence=0.5
                        )
                    except Exception as e:
                        st.warning(f"MediaPipe Model Init Error: {e}")


                while cap.isOpened():
                    ret, frame = cap.read()
                    if not ret:
                        st.write(f"Debug Info: 讀取結束或讀取失敗 (Frame {f_idx})")
                        break
                    f_idx += 1
                    # [v47 Fix] Dynamic Frame Interval
                    if f_idx % frame_interval != 0: continue
                    # [Fix] Reset annotated_frame to prevent stale images
                    if 'annotated_frame' in locals():
                        del annotated_frame

                    # 更新進度條 (避免過度更新拖慢速度，每 10 幀更新一次)
                    if f_idx % 10 == 0:
                        prog = min(f_idx / total_frames, 1.0)
                        st_progress.progress(prog)
                    try:
                        if use_mp and holistic_model:
                            # [v65 Major Upgrade] YOLO-Guided MediaPipe Stability Tracker with Specific ID Targeting
                            # 1. Use YOLO tracking to find IDs
                            tracker_file = create_tracker_config()
                            results_yolo = model.track(frame, persist=True, verbose=False, conf=0.20, iou=0.7, tracker=tracker_file, imgsz=480, classes=[0])

                            target_box = None
                            max_area = 0
                            found_specific_id = False

                            if results_yolo and len(results_yolo) > 0 and results_yolo[0].boxes is not None and results_yolo[0].boxes.id is not None:
                                ids = results_yolo[0].boxes.id.int().cpu().numpy()
                                boxes = results_yolo[0].boxes.xyxy.cpu().numpy()

                                for i, box in enumerate(boxes):
                                    track_id = ids[i]
                                    x1, y1, x2, y2 = box

                                    # If user selected a specific ID and we see it, lock immediately!
                                    if target_track_id > 0 and track_id == target_track_id:
                                        target_box = [int(x1), int(y1), int(x2), int(y2)]
                                        found_specific_id = True
                                        break

                                    # Fallback tracking: find the largest person if specific ID not found or target_track_id == 0
                                    area = (x2 - x1) * (y2 - y1)
                                    if area > max_area:
                                        max_area = area
                                        target_box = [int(x1), int(y1), int(x2), int(y2)]

                            annotated_frame = frame.copy()

                            if target_box is not None:
                                x1, y1, x2, y2 = target_box
                                h_frame, w_frame, _ = frame.shape

                                # Add 15% margin to the bounding box for full body context
                                bw = x2 - x1
                                bh = y2 - y1
                                mx = int(bw * 0.15)
                                my = int(bh * 0.15)

                                crop_x1 = max(0, x1 - mx)
                                crop_y1 = max(0, y1 - my)
                                crop_x2 = min(w_frame, x2 + mx)
                                crop_y2 = min(h_frame, y2 + my)

                                # Crop the image
                                roi = frame[crop_y1:crop_y2, crop_x1:crop_x2]

                                # 2. Feed ONLY the cropped largest person to MediaPipe
                                roi_rgb = cv2.cvtColor(roi, cv2.COLOR_BGR2RGB)
                                results_mp = holistic_model.process(roi_rgb)

                                if results_mp.pose_landmarks:
                                    # Function to translate normalized ROI coordinates back to original frame
                                    def translate_landmarks(landmark_list, crop_x1, crop_y1, roi_w, roi_h, frame_w, frame_h):
                                        if not landmark_list: return None
                                        # Create a deep copy using protobuf
                                        from google.protobuf.json_format import MessageToDict, ParseDict
                                        from mediapipe.framework.formats import landmark_pb2

                                        dict_lms = MessageToDict(landmark_list)
                                        new_landmarks = landmark_pb2.NormalizedLandmarkList()

                                        for i, lm in enumerate(landmark_list.landmark):
                                            # Convert to absolute ROI pixels
                                            abs_x = lm.x * roi_w
                                            abs_y = lm.y * roi_h
                                            # Add crop offset
                                            orig_abs_x = abs_x + crop_x1
                                            orig_abs_y = abs_y + crop_y1
                                            # Convert back to normalized original frame coordinates
                                            lm_dict = {'x': orig_abs_x / frame_w, 'y': orig_abs_y / frame_h, 'z': lm.z, 'visibility': lm.visibility}
                                            new_lm = new_landmarks.landmark.add()
                                            new_lm.x = lm_dict['x']
                                            new_lm.y = lm_dict['y']
                                            new_lm.z = lm_dict['z']
                                            new_lm.visibility = lm_dict['visibility']
                                        return new_landmarks

                                    roi_h, roi_w, _ = roi.shape

                                    # Translate all landmarks back to full frame
                                    global_pose = translate_landmarks(results_mp.pose_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)
                                    global_face = translate_landmarks(results_mp.face_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)
                                    global_lh = translate_landmarks(results_mp.left_hand_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)
                                    global_rh = translate_landmarks(results_mp.right_hand_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)

                                    # Draw on the original frame using translated landmarks
                                    mp_drawing.draw_landmarks(annotated_frame, global_pose, mp_holistic.POSE_CONNECTIONS, landmark_drawing_spec=mp_drawing_styles.get_default_pose_landmarks_style())
                                    if global_face:
                                        mp_drawing.draw_landmarks(annotated_frame, global_face, mp_holistic.FACEMESH_TESSELATION, None, mp_drawing_styles.get_default_face_mesh_tesselation_style())
                                    if global_lh:
                                        mp_drawing.draw_landmarks(annotated_frame, global_lh, mp_holistic.HAND_CONNECTIONS)
                                    if global_rh:
                                        mp_drawing.draw_landmarks(annotated_frame, global_rh, mp_holistic.HAND_CONNECTIONS)

                                    mid = target_track_id if found_specific_id else 1
                                    st.session_state.id_list.add(mid)
                                    st.session_state.display_mapping[mid] = mid

                                    if mid not in st.session_state.id_positions:
                                        st.session_state.id_positions[mid] = []
                                        st.session_state.id_motion_log[mid] = []
                                        st.session_state.id_tracking_count[mid] = 0

                                    st.session_state.id_tracking_count[mid] += 1

                                    lms = global_pose.landmark
                                    cx = int((lms[23].x + lms[24].x) * w_frame / 2) # Hip center for tracking
                                    cy = int((lms[23].y + lms[24].y) * h_frame / 2)
                                    st.session_state.id_positions[mid].append((f_idx, (cx, cy)))

                                    if mid not in st.session_state.id_actions:
                                        st.session_state.id_actions[mid] = {}
                                    st.session_state.id_actions[mid]["微觀追蹤(全身543點)"] = st.session_state.id_actions[mid].get("微觀追蹤(全身543點)", 0) + 1

                                    if mid not in st.session_state.id_features:
                                        st.session_state.id_features[mid] = {"clothing": "特定鎖定對象" if found_specific_id else "自動鎖定最大對象", "score_pending": True, "original_id": mid}

                                    # Use the YOLO bounding box for label
                                    nx, ny = crop_x1, max(0, crop_y1 - 20)
                                    label = f"ID: {mid} (Locked Target)" if found_specific_id else "ID: 1 (Largest Target)"
                                    t_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                                    cv2.rectangle(annotated_frame, (nx, max(0, ny - t_size[1] - 5)), (nx + t_size[0], ny), (0, 255, 0), -1)
                                    cv2.putText(annotated_frame, label, (nx, ny - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

                                    if len(st.session_state.id_positions[mid]) > 1:
                                        prev_pos = st.session_state.id_positions[mid][-2][1]
                                        dist = np.sqrt((cx-prev_pos[0])**2 + (cy-prev_pos[1])**2)
                                        st.session_state.id_motion_log[mid].append(dist)
                            else:
                                # Draw warning if no target found
                                cv2.putText(annotated_frame, "No Target Found", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

                            if 'st_frame' in locals() and annotated_frame is not None:
                                frame_rgb = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
                                st_frame.image(frame_rgb)
                            if out_video is not None:
                                out_video.write(annotated_frame)
                            continue # Skip YOLO processing

                        # [更新] 使用動態產生的設定檔
                        tracker_file = create_tracker_config()
                        # [PyInstaller Fix] Ensure tracker file exists or path is correct (created in current dir)

                        # [調整] 信心度門檻微調 (0.25 -> 0.20) 捕捉更多模糊 ID
                        # [調整] IoU (0.5 -> 0.7) 允許更多重疊 (Crowd Robustness)
                        # [v21] Cloud Optimization: imgsz=480
                        results = model.track(frame, persist=True, verbose=False, conf=0.20, iou=0.7, tracker=tracker_file, imgsz=480)

                        # 檢查是否有偵測到東西
                        if results and len(results) > 0 and results[0].boxes is not None and results[0].boxes.id is not None:
                            # [重要] 恢復彩色骨架：使用 plot() 但不畫預設框 (boxes=False)，保留骨架連線
                            try:
                                # [要求] 強制顯示彩色骨架連線 (kpt_line=True, kpt_radius=5)
                                annotated_frame = results[0].plot(boxes=False, labels=False, probs=False, kpt_line=True, kpt_radius=5)
                            except:
                                annotated_frame = frame.copy()

                            ids = results[0].boxes.id.int().cpu().numpy()
                            boxes = results[0].boxes.xyxy.cpu().numpy()
                            keypoints_data = results[0].keypoints.data.cpu().numpy() if results[0].keypoints is not None else None

                            # [Re-ID 步驟 1: 更新 Lost IDs]
                            current_mids = set([int(i) for i in ids])
                            # 找出已經穩定追蹤但本幀消失的 ID
                            active_mapped_ids = set()
                            for m in current_mids:
                                active_mapped_ids.add(st.session_state.id_map.get(m, m))

                            # 檢查 id_list 中有哪些人這幀不見了
                            for known_id in st.session_state.id_list:
                                if known_id not in active_mapped_ids:
                                    # 加入 Lost List (如果尚未加入)
                                    if known_id not in st.session_state.lost_ids and known_id in st.session_state.id_features:
                                        feat = st.session_state.id_features[known_id]
                                        # [v57 Fix] Save Last Position for Spatial Check
                                        last_pos = None
                                        if known_id in st.session_state.id_positions and st.session_state.id_positions[known_id]:
                                            last_pos = st.session_state.id_positions[known_id][-1][1] # (x,y)

                                        if 'hist' in feat:
                                            st.session_state.lost_ids[known_id] = {
                                                'hist': feat['hist'],
                                                'last_seen': f_idx,
                                                'last_pos': last_pos
                                            }

                            # 清理過期 Lost IDs (> 60 Frames)
                            keys_to_remove = [k for k, v in st.session_state.lost_ids.items() if f_idx - v['last_seen'] > 60]
                            for k in keys_to_remove:
                                del st.session_state.lost_ids[k]

                            # 顯示第一次成功的偵測
                            if not st.session_state.id_list and f_idx % 30 == 0:
                                pass

                            for i, box in enumerate(boxes):
                                mid = int(ids[i])
                                x1, y1, x2, y2 = map(int, box) # 提早提取座標
                                curr_center = ((x1+x2)//2, (y1+y2)//2)

                                # [v21.4 Logic Refactor] Removed early continue to allow all IDs to be mapped and tracked.
                                # This ensures the counting of IDs 1, 2, 3... remains consistent.
                                pass

                                # [Re-ID 步驟 2: 嘗試找回舊 ID]
                                # 只有當 mid 是新出現的 (不在 id_list) 且尚未被 map 過時才做
                                # 且必須有 lost_ids 才有意義
                                if mid not in st.session_state.id_list and mid not in st.session_state.id_map:
                                    if st.session_state.lost_ids:
                                        h = y2 - y1
                                        # 為了效能，只取中間區塊算直方圖 (避開背景)
                                        crop = frame[max(0, y1+int(h*0.1)):min(frame.shape[0], y2-int(h*0.1)), x1:x2]
                                        curr_hist = get_color_histogram(crop)

                                        best_match = -1
                                        best_score = 0.0 # 相關性最高為 1.0

                                        if curr_hist is not None:
                                            for lost_id, data in st.session_state.lost_ids.items():
                                                if data['hist'] is None: continue

                                                # [v57 Fix] Spatial Constraint (Distance Check)
                                                # Don't match if too far away (e.g. > 150px)
                                                if data['last_pos']:
                                                    dist = np.linalg.norm(np.array(curr_center) - np.array(data['last_pos']))
                                                    # Speed limit: e.g. 30px per frame? or just absolute
                                                    # If lost for 10 frames, max move is 300px.
                                                    # Let's use a reasonable bound for typical movement.
                                                    if dist > 200: continue 

                                                score = cv2.compareHist(curr_hist, data['hist'], cv2.HISTCMP_CORREL)
                                                # [v57 Fix] Stricter Threshold (0.65 -> 0.75) to prevent ID switching
                                                if score > 0.75 and score > best_score:
                                                    best_score = score
                                                    best_match = lost_id

                                            if best_match != -1:
                                                st.session_state.id_map[mid] = best_match
                                                if best_match in st.session_state.lost_ids:
                                                    del st.session_state.lost_ids[best_match] # 找回後從 lost 移除
                                                # st.write(f"Debug Re-ID: {mid} -> {best_match} (Score {best_score:.2f})")

                                # Apply Mapping (取出真實 ID)
                                mid = st.session_state.id_map.get(mid, mid)

                                # [ID 穩定性過濾]
                                st.session_state.id_tracking_count[mid] = st.session_state.id_tracking_count.get(mid, 0) + 1

                                # [v21.4 Fix] Correctly identify if this person is already the locked target
                                # We check the display_mapping to see what ID was previously assigned.
                                current_display_id = st.session_state.display_mapping.get(mid, -1)
                                is_target = (target_track_id > 0 and current_display_id == target_track_id) 
                                
                                if st.session_state.id_tracking_count.get(mid, 0) > 5 or is_target:
                                    # [v16] Map to Display ID (1..N)
                                    if mid not in st.session_state.display_mapping:
                                        st.session_state.display_mapping[mid] = len(st.session_state.display_mapping) + 1
                                    
                                    # Keep track of the raw mapped ID before swapping mid to Display ID
                                    raw_mapped_id = mid 
                                    mid = st.session_state.display_mapping[mid] # mid 現在是 1, 2, 3...

                                    # [v21.4 Fix] FILTER HERE
                                    if target_track_id > 0 and mid != target_track_id:
                                        continue

                                    st.session_state.id_list.add(mid)
                                    # x1, y1, x2, y2 已提取

                                    # [動作評分數據收集]
                                    center_x, center_y = (x1+x2)//2, (y1+y2)//2
                                    if mid not in st.session_state.id_positions:
                                        st.session_state.id_positions[mid] = []
                                        st.session_state.id_motion_log[mid] = []

                                    # 記錄 (frame_idx, (x,y)) 以便內插
                                    st.session_state.id_positions[mid].append((f_idx, (center_x, center_y)))

                                    # 計算單幀移動量作為 Amplitude 參考 (粗略)
                                    if len(st.session_state.id_positions[mid]) > 1:
                                        prev_pos = st.session_state.id_positions[mid][-2][1]
                                        curr_pos = (center_x, center_y)
                                        dist = np.sqrt((curr_pos[0]-prev_pos[0])**2 + (curr_pos[1]-prev_pos[1])**2)
                                        st.session_state.id_motion_log[mid].append(dist)

                                    # ID 1 (攝影師) 特別處理? 不，用動態評分即可解決 (攝影師不動 -> 1分)

                                    if mid not in st.session_state.id_features:
                                        # 上下半身分離 + HSV 顏色分析
                                        h = y2 - y1
                                        # 只取中間部分避免背景干擾
                                        shirt = frame[max(0, y1+int(h*0.1)):min(frame.shape[0], y1 + int(h*0.4)), x1+int((x2-x1)*0.2):x2-int((x2-x1)*0.2)]
                                        pants = frame[max(0, y1 + int(h*0.6)):min(frame.shape[0], y2-int(h*0.1)), x1+int((x2-x1)*0.2):x2-int((x2-x1)*0.2)]
                                        c_shirt = get_dominant_color(shirt)
                                        c_pants = get_dominant_color(pants)
                                        p_shirt = get_clothing_pattern(shirt) # 新增圖案分析

                                        # [修正] 確保下裝特徵也被考慮
                                        # 其實上衣跟下裝的邏輯是一樣的
                                        # 這裡只做簡單更新，避免覆蓋既有資訊? 不，我們每次都覆蓋最新的

                                        # 計算特徵字串
                                        # 注意：p_shirt 變數名可能混淆，我們直接用函式回傳

                                        feat_str = f"上衣：{c_shirt}{get_clothing_pattern(shirt)}。下裝：{c_pants}{get_clothing_pattern(pants)}、褲子。配件：無。"

                                        st.session_state.id_features[mid] = {
                                            "clothing": feat_str,
                                            "score_pending": True,
                                            "hist": get_color_histogram(shirt), # 儲存特徵供 Re-ID 使用
                                            "original_id": raw_mapped_id # [v16] 記錄原始 ID
                                        }

                                    # [v21 Restore] Action Recognition & Drawing Logic
                                    # [v21 Restore] Action Recognition & Drawing Logic
                                    color = (0, 140, 255) # Orange (Default)
                                    current_action = ""

                                    # Action Recognition
                                    if keypoints_data is not None and len(keypoints_data) > i:
                                        try:
                                            kpts = keypoints_data[i] # (17, 3)

                                            # [v42 Fix] Lying Down Detection

                                            # [v56 New] Multi-Action Detection Support
                                            current_actions = []

                                            # 1. Check Lying Down / Floor Action
                                            is_lying_down = False
                                            # Torso Orientation: Shoulders(5,6) vs Hips(11,12)
                                            if kpts[5][2]>0.5 and kpts[11][2]>0.5:
                                                dy = abs(kpts[5][1] - kpts[11][1])
                                                dx = abs(kpts[5][0] - kpts[11][0])
                                                if dy < dx * 0.8: # Horizontal
                                                    is_lying_down = True

                                            # Bounding Box Aspect Ratio (Backup)
                                            if not is_lying_down:
                                                w = x2 - x1
                                                h = y2 - y1
                                                if w > h * 1.2: 
                                                    is_lying_down = True

                                            if is_lying_down:
                                                current_actions.append("地板動作")
                                                color = (255, 0, 255) # Magenta

                                            # 2. Check Hands Up
                                            if (kpts[9][2] > 0.5 and kpts[5][2] > 0.5 and kpts[9][1] < kpts[5][1]) or \
                                            (kpts[10][2] > 0.5 and kpts[6][2] > 0.5 and kpts[10][1] < kpts[6][1]):
                                                current_actions.append("舉手")
                                                if not is_lying_down: color = (0, 0, 255) # Red priority if standing

                                            # 3. Check Squat
                                            if not is_lying_down:
                                                if kpts[11][2]>0.5 and kpts[13][2]>0.5 and kpts[14][2]>0.5:
                                                    thigh_len_l = abs(kpts[13][1] - kpts[11][1])
                                                    shin_len_l = abs(kpts[15][1] - kpts[13][1]) if kpts[15][2]>0.5 else 100
                                                    if thigh_len_l < shin_len_l * 0.5: 
                                                        current_actions.append("蹲下")
                                                        color = (255, 0, 0) # Blue

                                            # 4. Check Jump
                                            if not is_lying_down and mid in st.session_state.id_positions:
                                                hist = st.session_state.id_positions[mid]
                                                if len(hist) > 5:
                                                    recent_hist = hist[-15:] 
                                                    ys = [p[1][1] for p in recent_hist]
                                                    avg_y = np.mean(ys)
                                                    h = y2 - y1
                                                    if center_y < avg_y - (h * 0.20): 
                                                        current_actions.append("跳躍")
                                                        color = (255, 255, 0) # Cyan

                                            # 5. Check Kick
                                            if not is_lying_down and not "蹲下" in current_actions:
                                                if kpts[15][2]>0.5 and kpts[13][2]>0.5:
                                                    if kpts[15][1] < kpts[13][1]:
                                                        current_actions.append("踢腿/抬腿")
                                                elif kpts[16][2]>0.5 and kpts[14][2]>0.5:
                                                    if kpts[16][1] < kpts[14][1]:
                                                        current_actions.append("踢腿/抬腿")

                                            # 6. Check Running
                                            if not is_lying_down and mid in st.session_state.id_positions:
                                                hist = st.session_state.id_positions[mid]
                                                if len(hist) > 3:
                                                    prev_x = hist[-3][1][0]
                                                    curr_x = center_x
                                                    dx_speed = abs(curr_x - prev_x)
                                                    w = x2 - x1
                                                    if dx_speed > w * 0.5:
                                                        current_actions.append("跑步")

                                            # Consolidate Actions
                                            current_actions = list(set(current_actions))
                                            display_action_en = ""

                                            if current_actions:
                                                action_map = {
                                                    "舉手": "HandsUp", "蹲下": "Squat", "跳躍": "Jump",
                                                    "地板動作": "Floor", "踢腿/抬腿": "Kick", "跑步": "Run"
                                                }
                                                english_tags = [action_map.get(a, a) for a in current_actions]
                                                display_action_en = "+".join(english_tags)
                                            else:
                                                display_action_en = ""

                                            # Drawing
                                            # [v21.3 Refine] Thicker boxes (2->4) and larger font (0.6->0.8)
                                            cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), color, 4)
                                            
                                            # [v21.3 Fix] Always show label with ID, even if no action
                                            label = f"ID: {mid}"
                                            if display_action_en:
                                                label += f" {display_action_en}"
                                            
                                            t_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 3)[0]
                                            c2 = x1 + t_size[0], y1 - t_size[1] - 10
                                            cv2.rectangle(annotated_frame, (x1, y1), (c2[0], c2[1]), color, -1)
                                            cv2.putText(annotated_frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 3)

                                            # Logging
                                            if current_actions:
                                                if mid not in st.session_state.id_actions:
                                                    st.session_state.id_actions[mid] = {}
                                                for act in current_actions:
                                                    st.session_state.id_actions[mid][act] = st.session_state.id_actions[mid].get(act, 0) + 1

                                        except Exception as e:
                                            pass

                                    # (2) Focus Analysis (Head Yaw)
                                    head_yaw = 0
                                    if keypoints_data is not None and len(keypoints_data) > i:
                                        try:
                                            kpts = keypoints_data[i]
                                            nose = kpts[0][:2]
                                            l_ear = kpts[3][:2]
                                            r_ear = kpts[4][:2]
                                            if kpts[0][2] > 0.5 and kpts[3][2] > 0.5 and kpts[4][2] > 0.5:
                                                head_yaw = calculate_head_yaw(nose, l_ear, r_ear)
                                        except:
                                            pass

                                    if mid not in st.session_state.id_yaw_history:
                                        st.session_state.id_yaw_history[mid] = []
                                    st.session_state.id_yaw_history[mid].append(head_yaw)

                                    # (3) Social Interaction
                                    my_center = (center_x, center_y)
                                    for other_mid, pos_list in st.session_state.id_positions.items():
                                        if other_mid == mid: continue
                                        if not pos_list: continue
                                        last_frame_idx, other_pos = pos_list[-1]
                                        if abs(last_frame_idx - f_idx) > 5: continue
                                        dist = np.linalg.norm(np.array(my_center) - np.array(other_pos))
                                        if dist < 150:
                                            pair = tuple(sorted((mid, other_mid)))
                                            st.session_state.id_interactions[pair] = st.session_state.id_interactions.get(pair, 0) + 1

                        # [Fix] Update UI (Outside Loop)
                        if 'annotated_frame' not in locals():
                            annotated_frame = frame

                        if 'st_frame' in locals() and annotated_frame is not None:
                            frame_rgb = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
                            st_frame.image(frame_rgb)

                        # [v21.4 Fix] ALWAYS write frame to output video, even if no detection
                        # This ensures the output video has the correct length and sync.
                        if out_video is not None:
                            try:
                                out_video.write(annotated_frame if annotated_frame is not None else frame)
                            except:
                                pass

                    except Exception as e:
                        if "lap" in str(e) or "LAP" in str(e):
                            st.error("❌ 缺少 'lap' 模組。請確認 requirements.txt 包含 'lapx'。")
                            st.stop()
                        pass

                cap.release()
                if 'holistic_model' in locals() and holistic_model:
                    holistic_model.close()
                if out_video is not None and out_video.isOpened():
                    out_video.release() # Release writer
                    
                # [v21.5.1 Hotfix] Move conversion BEFORE rerun to avoid race condition
                if os.path.exists(output_path):
                    converted_path = os.path.abspath("obs_video_h264.mp4")
                    st.info("🔄 正在轉換影片格式 (H.264) 以支援網頁播放...")
                    
                    import shutil
                    import subprocess
                    if shutil.which("ffmpeg"):
                        try:
                            # Use ultrafast and crf 28 for cloud performance
                            cmd = [
                                "ffmpeg", "-y", "-i", output_path,
                                "-c:v", "libx264", "-preset", "ultrafast", "-crf", "28",
                                "-f", "mp4", converted_path
                            ]
                            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=90)
                            if os.path.exists(converted_path) and os.path.getsize(converted_path) > 100:
                                st.session_state.video_output_path = converted_path
                                logging.info(f"FFmpeg Success: {converted_path}")
                            else:
                                st.session_state.video_output_path = output_path
                        except Exception as e:
                            logging.warning(f"FFmpeg conversion failed: {e}")
                            st.session_state.video_output_path = output_path
                    else:
                        st.session_state.video_output_path = output_path
                else:
                    st.session_state.video_output_path = None

                st_progress.empty() # 清除進度條

                # [v21.5.2 Fix] Remove redundant path overwriting that happened in v21.5.1
                # The video_output_path is already set by the FFmpeg block above.
                if 'video_output_path' not in st.session_state or st.session_state.video_output_path is None:
                    if out_video is not None:
                         st.session_state.video_output_path = output_path

                # [v66] 直接在此處完成所有數據結算，避免二次 Rerun 失敗
                logging.info("Calculating Group Sync (Kuramoto)...")
                group_sync_r = calculate_kuramoto_order_parameter(st.session_state.id_motion_log)
                st.session_state.group_sync_r = group_sync_r if group_sync_r is not None else 0.0

                st.session_state.final_id_count = len(st.session_state.id_list)
                st.session_state.final_id_list = sorted(list(st.session_state.id_list))

                st.success("✅ 分析完成！正在跳轉報表...")
                st.session_state.analysis_done = True
                st.rerun() 

            except Exception as e:
                st.error(f"發生系統錯誤: {e}")
                st.write(traceback.format_exc())
            finally:
                # 清理暫存檔
                if os.path.exists(tfile_path):
                    try:
                        # os.remove(tfile_path) # Debug: Keep for now
                        pass 
                    except:
                        pass

                # [v21.5.1 Fix] Clean up legacy conversion block from finally
                pass


# [v66] The transition results block has been consolidated into the analysis completion block above.
# The code below now handles the final report rendering.


# [v65 Fix] Stop execution only if analysis is NOT complete
# This prevents the report below from prematurely rendering on page load
# but allows it to render after st.rerun() when analysis_done is True.
if not st.session_state.analysis_done:
    st.stop()

# --- 4. 報表編輯區 (分析完成後鎖定顯示) ---

# [新] 持續顯示 ID 資訊 (使用者要求保留在上方)
if 'final_id_count' in st.session_state and st.session_state.final_id_count > 0:
    st.info(f"📊 偵測到的 ID 數量: {st.session_state.final_id_count}")
    st.write(f"ID 列表: {st.session_state.final_id_list}")
    if 'group_sync_r' in st.session_state:
        st.write(f"🔗 群體同步率 (R值): {st.session_state.group_sync_r} (1.0 為完全同步)")

if st.session_state.last_frame is not None:
    st.image(st.session_state.last_frame, caption="📌 最終偵測畫面 (請根據此畫面 ID 填寫下方報表)")

st.markdown("---")
# [v17.1 Fix] Replay Button Below Video Analysis
# 使用者要求：按鈕在下方，點擊後展開播放器，可重複觀看
st.markdown("---")
# [v18.7 Fix] Always show video after analysis
if 'video_output_path' in st.session_state and os.path.exists(st.session_state.video_output_path):
    st.info(f"🎥 鑑識影片回放 ({st.session_state.video_output_path})")
    st.video(st.session_state.video_output_path)

    # [v18.8 Fix] Download Button
    with open(st.session_state.video_output_path, "rb") as f:
        st.download_button(
            label=" 下載影片 (Download Video)",
            data=f,
            file_name="analysis_video.mp4",
            mime="video/mp4",
            key="download_video_btn"
        )

st.markdown("---")
st.subheader("📊 第二步：HMEAYC 專業觀察編輯")

s_ids = sorted(list(st.session_state.id_list))
if not s_ids:
    if target_track_id > 0:
        st.warning(f"⚠️ 分析完成，但在影片中未抓取到您指定的 ID {target_track_id}。")
        st.info(f"💡 目前影片中總共偵測到 {len(st.session_state.display_mapping)} 位不同的幼兒。如果您想觀察其他孩子，請調整左側的追蹤 ID。")
    else:
        st.warning("⚠️ 偵測過程中未抓取到有效 ID，請重新上傳清晰影片。")
else:
    # [v11 Update] 選擇教師 ID (用於計算師生同步率)
    # 製作選項列表: "ID_1 (原:80)"
    id_options = []
    id_map_rev = {} # 顯示名稱 -> 真實 ID

    for idx, m in enumerate(s_ids, 1):
        # [v40 Fix] Use Raw ID directly to match Video Overlay
        # User confusion: Video says ID 9, Table says ID 7 (Original 9).
        # Solution: Table should say ID 9.
        label = f"ID_{m}"
        id_options.append(label)
        id_map_rev[label] = m

    col_t1, col_t2 = st.columns([1, 3])
    with col_t1:
        teacher_label = st.selectbox("請選擇教師 (示範者) ID:", ["無"] + id_options)

    teacher_id = None
    if teacher_label != "無":
        teacher_id = id_map_rev[teacher_label]
        st.info(f"已設定 {teacher_label} 為教師，將計算其他幼兒與其的動作同步率。")

    # 從 Session State 讀取資料
    df_list = []

    # [v21] Pre-Calc Removed (Using Two-Pass)

    for idx, m in enumerate(s_ids, 1): # idx 從 1 開始
        # [v8 Fix] 先取出 feat 才能判斷 original_id
        feat = st.session_state.id_features.get(m, {"clothing": "分析中"})
        if isinstance(feat, str): feat = {"clothing": feat}

        # [v40 Fix] Use Raw ID for Label
        original_id = m # We rely on m directly now
        # if "original_id" in feat: original_id = feat["original_id"] # Legacy logic ignored for clarity

        label = f"ID_{m}" # Consistent Label ID_9

        # 計算動態評分
        pos_history = st.session_state.id_positions.get(m, [])
        # 解壓縮 (frame, pos) -> 只取 pos
        pos_only = [p[1] for p in pos_history]
        score = get_motion_score(pos_only) 

        # [v56 Enh] Metric 1: Height-Normalized Motion Energy
        # Need to get average body height for this ID to normalize
        # st.session_state.id_positions[mid] has (frame, (x,y)) but we need (x,y,w,h) or just look up bbox history if we stored it?
        # We didn't store BBox history efficiently.
        # Alternative: Use the "Body Length" (Neck to Hip) from keypoints?
        # Workaround: We have "energy" calculated from raw pixel movement.
        # Let's use a heuristic: Assuming 480p height. 
        # Better: We can estimate body scale from the last checking frame or average.
        # But simpler is just to refine the raw energy scale based on typical values.
        # Let's stick to the current "energy" but map it better.
        # Current energy is mean of pixel displacement. 
        # 1-3 is low, 3-6 is med, >6 is high.
        # Let's adjust rounding.
        energy = 0
        if m in st.session_state.id_motion_log:
            if st.session_state.id_motion_log[m]:
                raw_energy = np.mean(st.session_state.id_motion_log[m])
                # Normalize: assume typical body height is ~150px in 480p (approx).
                # So 1% movement is 1.5px.
                # Energy 5px -> 3.3% body size.
                # Let's map 0-15px to 1-10 score.
                energy = min(round(raw_energy * 0.8, 1), 10.0) # slightly scaled down

        # [v56 Enh] Metric 2: Smoothness (Stability) using Jerk
        # Smoothness = 1.0 / (1.0 + Mean_Jerk) * 100
        # We need standard deviation of acceleration?
        # Let's use the existing helper 'calculate_smoothness' or simple variance of motion
        smoothness_score = 0
        if m in st.session_state.id_motion_log:
            m_log = st.session_state.id_motion_log[m]
            if len(m_log) > 2:
                # Calculate Acceleration (diff of speed)
                accels = np.diff(m_log)
                # Calculate Jerk (diff of accel)
                jerks = np.diff(accels)
                input_jerk = np.mean(np.abs(jerks))
                # Map Jerk (0.0 to 5.0 typically) to 0-100 Stability
                # Low Jerk = High Stability
                # 0.0 -> 100
                # 5.0 -> 0 (Shakey)
                # Formula: 100 * exp(-k * jerk)
                # [v62 Refine] Coefficient tuned to -0.08 for realistic score spread (User Request: "不要太寬，希望真實評分")
                smoothness_score = round(100 * np.exp(-0.08 * input_jerk), 1)

        # [v11 New] 計算師生同步率
        sync_score = None
        if teacher_id is not None:
            # [Fix] Robust comparison for Teacher ID (Sync = 100%)
            # 使用 int 比較避免 string/np.int64 誤差
            try:
                if int(m) == int(teacher_id):
                    sync_score = 100.0 # 老師本人
                else:
                    teacher_pos = st.session_state.id_positions.get(teacher_id, [])
                    logging.info(f"Debug Sync: Student {m} vs Teacher {teacher_id}")
                    # [v18 Fix] Pearson Correlation returns -1 to 1. Scale to percentage.
                    raw_r = calculate_teacher_sync(pos_history, teacher_pos)

                    if raw_r is not None:
                        # [v18.8 Fix] Safety scaling: if score is <= 1.0 (e.g. 0.8), multiply by 100 to get 80.0
                        # This handles cases where function returns 0-1 instead of 0-100
                        sync_score = float(raw_r)
                        if 0 < sync_score <= 1.0:
                            sync_score *= 100
                        sync_score = round(sync_score, 1)
                        # logging.info(f"Debug Sync: Auto-scaled {raw_r} to {sync_score}")
                    else:
                        sync_score = 0.0

                    if teacher_id is not None and m != teacher_id:
                        # Use motion logs (energy) for cross-correlation
                        student_log = st.session_state.id_motion_log.get(m, [])
                        teacher_log = st.session_state.id_motion_log.get(teacher_id, [])
                        # [v47 Fix] Use effective FPS for accurate lag time
                        eff_fps = st.session_state.get("effective_fps", 30)
                        temp_corr, lag_sec = analyze_temporal_sync(student_log, teacher_log, fps=eff_fps)

                    # [v26 Add] Store for Table
                    feat['temp_lag'] = lag_sec
                    feat['temp_corr'] = temp_corr

            except Exception as e:
                logging.error(f"Sync error for {m}: {e}")
                sync_score = 0.0

        # [v18.9 Fix] Final Hard Override for Teacher
        # If for ANY reason the logic failed, force it now.
        if teacher_id is not None:
            try:
                if int(m) == int(teacher_id):
                    sync_score = 100.0
            except:
                pass

        # 特徵補強 (v12: 填入動作與視線)
        # 篩選出現頻率高的動作 (例如 > 20% frames)
        # 這裡簡單起見，只要出現次數 > 10 Frames 就視為有效
        feat_enhance = ""
        action_counts = st.session_state.id_actions.get(m, {})
        valid_tags = []

        # 優先級 sorting
        priority = ["跳躍", "地板動作", "舉手", "蹲下", "踢腿/抬腿", "跑步", "側臉", "專注"] 

        for tag in priority:
            if action_counts.get(tag, 0) > 3: # [v28 Fix] Lower threshold from 10 to 3 frames
                valid_tags.append(tag)

        if valid_tags:
            feat_enhance = ", ".join(valid_tags)

        # [v14] 生成 AI 評語
        gaze_status = "一般"
        if "專注" in valid_tags: gaze_status = "專注"
        if "側臉" in valid_tags: gaze_status = "側臉"

        # [v19 New] Calculate Focus Score (Gaze at Teacher)
        focus_score = 0
        if teacher_id is not None and m != teacher_id and teacher_id in st.session_state.id_positions:
            t_pos_list = st.session_state.id_positions[teacher_id]
            if t_pos_list:
                # Teacher Avg Pos (Simplified: Last known)
                t_avg_x = t_pos_list[-1][1][0]

                yaws = st.session_state.id_yaw_history.get(m, [])
                s_positions = st.session_state.id_positions.get(m, [])

                # Check focus
                valid_f = 0
                focused_f = 0
                min_len = min(len(yaws), len(s_positions))

                for i in range(0, min_len, 5): # Sample every 5 frames
                    s_x = s_positions[i][1][0]
                    yaw = yaws[i]
                    if check_gaze_at_target((s_x, 0), yaw, (t_avg_x, 0)):
                        focused_f += 1
                    valid_f += 1

                if valid_f > 0:
                    focus_score = int((focused_f / valid_f) * 100)

        # [v19.2 Refine] Social Role Logic
        interaction_count = 0
        if 'id_interactions' in st.session_state:
            interaction_count = sum([c for (pair, c) in st.session_state.id_interactions.items() if m in pair])

        # Default
        role = "獨立觀察 (Independent)" 

        # Hierarchy of Roles
        if interaction_count >= 30 and score >= 3:
            role = "社交活躍 (Active)"
        elif interaction_count >= 30:
            role = "靜態互動 (Passive)"
        elif focus_score >= 60:
            role = "專注跟隨 (Focused)"
        elif sync_score is not None and sync_score < 40 and energy > 5.0:
            role = "自主探索 (Creative)"
        elif sync_score is not None and sync_score >= 80:
            role = "動作模仿 (Imitating)"

        # [Fix] Teacher is always Teacher
        if teacher_id is not None and m == teacher_id:
            role = "教學者 (Teacher)"

        # [v65 New] Micro-Analysis Override
        # Since Micro mode isolates one subject, group/social metrics are disabled
        if "MediaPipe" in perf_mode:
            role = "微觀特寫 (Micro-Analysis)"
            sync_score = None # Disable sync comparison
            # focus_score remains relevant if they look at camera/forward, but if no teacher exists, it stays 0.

        # 取出動作 (排除 gaze 標籤)
        pure_actions = [tag for tag in valid_tags if tag not in ["專注", "側臉"]]

        # [v56 New] 7-Type Rich Text AI Comments Logic
        ai_comment = ""

        # Define Context
        is_teacher = (teacher_id is not None and m == teacher_id)
        sync_val = sync_score if sync_score is not None else 0
        energy_val = energy # 0-10
        stab_val = smoothness_score # 0-100
        inter_val = interaction_count

        # 1. 🌟 Role Model (小老師)
        if role == "微觀特寫 (Micro-Analysis)":
            # [v65 New] Micro-Analysis Specific Comments
            if stab_val > 70 and energy_val > 5.0:
                options = [
                    "動作執行度極高，肢體延展充分且控制精準，展現出高度的肌肉協調性。",
                    "肢體控制展現專業水準，動作核心穩定且張力十足，反映出絕佳的平衡感。",
                    "動作細節處理非常細膩，各關節的配合展現出高度的律動天分與穩定度。",
                    "能精準控制每一次的肢體擺動，重心變換流暢，展現出高度的學習穩定性。",
                    "具備卓越的關節活動度與控制力，動作優雅且精確度極高。"
                ]
            elif stab_val > 70:
                options = [
                    "動作重心相當穩定，雖然活動幅度不大，但對肢體的細微控制展現出沉穩特質。",
                    "重心控制極佳，體現出優良的穩定性，對於細節動作的觀察相當入微。",
                    "動作穩定且從容，表現出專注的身體感知能力，能精確掌握微小的肢體變化。",
                    "表現出沉著的穩定性，即使是靜態動作也展現出良好的肌肉支撐與耐力。",
                    "穩定性出色，重心分配平均，能以安靜且精確的方式內化動作要領。"
                ]
            elif energy_val > 7.0:
                options = [
                    "展現出強大的爆發能量，各關節的大幅度變化顯示其高度的參與動機與爆發力。",
                    "能量充沛，動作力度十足且充滿張力，充分展現了肢體的生命力與自信。",
                    "各部位肌肉動員能力極佳，動作果斷有力，具備強大的節奏穿透力。",
                    "活力十足，能在極短時間內完成大幅度的動作轉換，爆發力相當優異。",
                    "動作充滿力量感，對空間的探索極致且大膽，展現出極強的表現慾。"
                ]
            else:
                options = [
                    "目前呈現較為基礎的律動節奏，骨架變化反映出正在逐步適應肢體伸展的過程。",
                    "動作節奏正在發展中，肢體伸展呈現出初步的協調性，進步空間值得期待。",
                    "正處於對肢體掌控的探索階段，建議透過更多元化的小幅度動作練習來建立信心。",
                    "表現出對律動的基本反應，隨著訓練持續，肢體的延展度與流暢性將會顯著提升。",
                    "律動感尚屬基礎，但各部位的配合已初具雛形，是一個良好的學習起點。"
                ]
            ai_comment = random.choice(options)
        elif is_teacher or (sync_val > 80 and stab_val > 70):
            options = [
                "動作精準且穩定，完全跟隨老師指令，展現出極佳的專注力與示範能力。",
                "✅ 完美的小老師！動作與指令高度一致，是大家的學習榜樣。",
                "🌟 展現出超齡的穩定性與節奏感，動作執行度滿分。",
                "動作細節處理到位，肢體控制力極佳，展現出穩定的領導氣質。",
                "專注力極高，對指令的反應迅速且精確，展現出卓越的學習成效。"
            ]
            ai_comment = random.choice(options)

        # 2. 🚀 Energetic Mover (活力充沛)
        elif energy_val > 7.0 and sync_val > 40:
            options = [
                "充滿活力，動作幅度大且熱情，能帶動周圍氣氛，但偶爾需注意細節控制。",
                "⚡ 動作能量充沛，全身心投入活動，展現出高度的參與熱情。",
                "🚀 活動量大，肢體伸展充分，若能稍微收斂動作範圍會更精準。",
                "展現出強大的爆發力與熱情，每一次擺動都充滿生命力。",
                "活力四射，大幅度的肢體動作顯示出對活動的高度興奮與投入。"
            ]
            ai_comment = random.choice(options)

        # 3. 🐢 Observer (冷靜觀察者)
        elif energy_val < 3.0 and focus_score > 60:
            options = [
                "雖動作較少，但視線持續專注於示範者，屬於「觀察型」學習風格，正在內化動作。",
                "🐢 動作偏靜態，但注意力高度集中，可能正在仔細觀察老師的每一個細節。",
                "👀 雖然外在動作不大，但專注度極高，展現出沉穩的學習態度。",
                "靜靜地站在位置上，但眼神始終跟隨著指示，展現出深度的思考與觀察。",
                "屬於先觀察後行動的學習者，正在安靜地吸收所有動作訊息。"
            ]
            ai_comment = random.choice(options)

        # 4. 🎨 Creative Explorer (自主探索) / Low Sync in Imitation
        elif sync_val < 40 and energy_val > 5.0:
            options = [
                "🎨 展現出獨特的動作詮釋，不拘泥於模仿，擁有豐富的自主探索與創造力。",
                "✨ 動作充滿創意與變化，似乎正在進行個人的肢體探索，展現獨特思維。",
                "🌈 跳脫了制式動作的框架，展現出極具個人特色的律動風格。",
                "擁有豐富的想像力，將音樂與動作轉化為個人獨有的肢體語言。",
                "不被模仿所限制，正以直覺探索身體的各種可能性，展現高度原創性。"
            ]
            ai_comment = random.choice(options)

        # 5. 🤝 Social Butterfly (社交活躍)
        elif inter_val > 20 and sync_val > 40:
            options = [
                "🤝 與同儕互動頻繁，樂於分享與模仿同伴動作，透過人際互動來學習。",
                "🦋 在團體中互動活躍，比起跟隨老師，更傾向與身邊同學建立連結。",
                "👥 展現出高度的社交意願，在互動中尋找動作的節奏與樂趣。",
                "樂於在群體中移動，透過觀察與回應同伴의 動作來提升參與感。",
                "是一個天生的社交者，動作中常帶著對周圍夥伴的關注與互動。"
            ]
            # [v21.4 Fix] High-Entropy randomization for AI comments
            # Use child's ID, current time, and a unique hash of their motion history
            seed_val = int(time.time() * 1000) + m + len(st.session_state.id_positions.get(m, []))
            rand_gen = random.Random(seed_val)
            ai_comment = rand_gen.choice(options)

        # 6. 🌪️ Active Wanderer (探索漫遊) - Distracted
        elif focus_score < 30 and energy_val > 5.0:
            options = [
                "🌪️ 精力旺盛但較難聚焦，動作轉換頻繁，建議透過明確指令引導其注意力。",
                "⚡ 活動力強但略顯分心，容易被周圍環境吸引，需要更多的眼神接觸引導。",
                "🎢 肢體動作豐富但缺乏方向，建議給予具體任務以收斂其注意力。",
                "在空間中自由探索，雖然活動量足夠，但需要引導核心目標。",
                "容易受外部環境干擾，動作較為跳躍，需建立更明確的視覺焦點。"
            ]
            ai_comment = random.choice(options)

        # 7. 🧘 Steady Follower (穩定跟隨) - Default Good
        else:
            options = [
                "🧘 動作平穩流暢，依照自己的節奏跟隨，展現出良好的身體控制能力。",
                "✅ 穩定地參與活動，動作不疾不虛，能跟上團體的大致節奏。",
                "🍃 動作表現平穩，雖然沒有特別突出，但始終保持在團體節奏中。",
                "按部就班地執行動作，展現出踏實的學習態度與基本的律動感。",
                "能在輕鬆的狀態下保持參與，其平穩的節奏感有助於穩定身心。"
            ]
            ai_comment = random.choice(options)

        # [v21 Upgrade] Phase 1: Store Raw Data
        lag_display = "-"
        if feat.get('temp_lag', 0) and abs(feat.get('temp_lag', 0)) > 0.1:
            lag_display = f"{feat.get('temp_lag', 0):.2f}s"

        # [v56 cleanup] Old logic removed.

        # [v61 Fix] Generate Combined Context-Aware AI Comment
        final_comment = generate_expert_comment(score, sync_score, focus_score, role, pure_actions, class_stats=None, archetype_text=ai_comment)

        df_list.append({
            "序號": idx, 
            "幼兒 ID": f"ID_{idx} (原:{original_id})", 
            "AI 服裝特徵": feat.get("clothing", ""),
            "特徵補強 (圖案/熊/亮片)": None, 
            "AI 觀察判定 (1-5)": score,
            "跟隨指令 (同步率%)": float(f"{sync_score:.0f}") if sync_score is not None else 0, 
            "時序延遲 (Lag)": lag_display, 
            "專注度(%)": focus_score, 
            "參與型態": role,        
            "動作檢測 (舉手、側臉)": ", ".join(pure_actions), # [v56 Fix] Use pure_actions 
            "AI 總結評語": final_comment, # [v61 Fix] Use Combined Comment

            "教師評分 (1-5)": None,
            "教師評語": None,
            "動作能量": energy, # [New]
            "動作穩定度": smoothness_score, # [v56 New]
        })

    # 轉換為 DataFrame 
    df = pd.DataFrame(df_list)

    # [v20.2 Sort] Sort by ID number
    if not df.empty:
        df['sort_key'] = df['幼兒 ID'].apply(lambda x: int(x.split(' ')[0].split('_')[1]) if '_' in x else 999)
        df = df.sort_values('sort_key').drop(columns=['sort_key'])

    # [v65 New] Custom Column Layout for Micro-Analysis
    if "MediaPipe" in st.session_state.get('perf_mode', ''):
        required_cols = ["序號", "幼兒 ID", "AI 服裝特徵", "特徵補強 (圖案/熊/亮片)", "AI 觀察判定 (1-5)", "時序延遲 (Lag)", "專注度(%)", "動作能量", "動作穩定度", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評分 (1-5)", "教師評語"]
    else:
        required_cols = ["序號", "幼兒 ID", "AI 服裝特徵", "特徵補強 (圖案/熊/亮片)", "AI 觀察判定 (1-5)", "跟隨指令 (同步率%)", "時序延遲 (Lag)", "專注度(%)", "動作能量", "動作穩定度", "參與型態", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評分 (1-5)", "教師評語"]

    if df.empty:
        df = pd.DataFrame(columns=required_cols)
    else:
        for col in required_cols:
            if col not in df.columns:
                df[col] = None 
        df = df.reindex(columns=required_cols) # Ensures exact column order and drops unwanted columns

    # [v31 Fix] Persistent Naming Logic using Session State
    if 'custom_name_map' not in st.session_state:
        st.session_state.custom_name_map = {}

    # Preserve the Original Raw ID for mapping (Hidden Column)
    # We need a column that stays constant even if "幼兒 ID" is edited.
    # "幼兒 ID" acts as the Display/Edit column.
    # "Raw_ID" acts as the Key.

    # 1. Inject Raw_ID for tracking
    # 1. Inject Raw_ID for tracking
    if not df.empty:
        # Re-extract raw ID from the "幼兒 ID" string if needed, or assume it's unique enough
        # Current "ID_X (原:Y)" is unique per session run (until restart)
        # [v45 Fix] Extract Simple Key "ID_X" for consistent mapping
        # This ensures that even if display name is complex, the key remains stable.
        df['Raw_ID'] = df['幼兒 ID'].apply(lambda x: x.split(" ")[0]) if not df.empty else []

        # 2. Apply existing map to Display Column
        # If Raw_ID is in map, update "幼兒 ID" to show the Custom Name
        # [v45 Fix] Logic: If key exists, use it; else keep original display (w/ suffix)
        def apply_name(row):
            key = row['Raw_ID']
            return st.session_state.custom_name_map.get(key, row['幼兒 ID'])

        df['幼兒 ID'] = df.apply(apply_name, axis=1)

    # 設定欄位格式
    # [v19 Fix] Restore st.data_editor call
    # [v36 Fix] Define Callback for Persistent Renaming
    def update_names_callback():
        """
        Callback to handle name changes immediately before rerun.
        Uses raw index to map back to ID key because data_editor uses 0-based index.
        """
        # Access the editor state directly
        editor_state = st.session_state.get("data_editor_v31_final", {})
        edited_rows = editor_state.get("edited_rows", {})

        if not edited_rows:
            return

        # Reconstruct the ID list to find the key
        # Must match the sort order used in DF construction
        if 'final_id_list' in st.session_state:
            s_ids = st.session_state.final_id_list
        else:
            s_ids = sorted(list(st.session_state.id_list))

        # [Fix] Missing Loop restored
        for idx, changes in edited_rows.items():
            if "幼兒 ID" in changes:
                new_name = changes["幼兒 ID"]
                # Find the Raw Key (e.g. "ID_9")
                try:
                    # idx from data_editor is the row index in the displayed DF
                    m = s_ids[int(idx)] 
                    raw_key = f"ID_{m}"

                    # Update Map
                    st.session_state.custom_name_map[raw_key] = new_name
                except IndexError:
                    pass # Should not happen if sync is correct

    edited_df = st.data_editor(
        df, 
        use_container_width=True,
        column_config={
            "Raw_ID": None, # Hide the boolean/key column
            "序號": st.column_config.NumberColumn("序號", format="%d", width=40, disabled=True), 
            "動作能量": st.column_config.ProgressColumn(
        "動作能量 (Energy)", 
        help="動作的大小與劇烈程度 (0-100)", 
        min_value=0, max_value=100,
        format="%d"
    ),
    "動作穩定度": st.column_config.ProgressColumn(
        "動作穩定度 (Smoothness)", 
        help="[Neuro-Motor] 動作流暢性，低分代表抖動或不穩 (0-100)", 
        min_value=0, max_value=100,
        format="%d"
    ),
    "專注度(%)": st.column_config.ProgressColumn("專注度", min_value=0, max_value=100, format="%d%%", width=80), 
            "參與型態": st.column_config.TextColumn("參與型態", width=120), 
            "幼兒 ID": st.column_config.TextColumn( 
                "幼兒 ID (可修改姓名)", 
                width=150, 
                disabled=False,
                help="點擊兩下修改姓名，系統會自動記憶 (直到重整網頁)"
            ),
            "AI 服裝特徵": st.column_config.TextColumn("AI 服裝特徵", width=300), 
            "特徵補強 (圖案/熊/亮片)": st.column_config.TextColumn("特徵補強", width=100), 
            "AI 觀察判定 (1-5)": st.column_config.NumberColumn("AI 評分", width=80), 
            "跟隨指令 (同步率%)": st.column_config.NumberColumn("同步率", format="%.0f", width=80), 
            "動作檢測 (舉手、側臉)": st.column_config.TextColumn("動作檢測", width=150), 
            "AI 總結評語": st.column_config.TextColumn("AI 總結評語", width=600), 
            "教師評分 (1-5)": st.column_config.NumberColumn("教師評分", width=80), 
            "教師評語": st.column_config.TextColumn("教師評語", width=200),
        },
        hide_index=True,
        key="data_editor_v31_final", # Unique key
        on_change=update_names_callback # [v36] Bind callback
    )

    # [v36] Removed old manual diff logic (lines 2077-2100) as callback handles it robustly.
    # Check if we need to force rerun?
    # Streamlit automatically reruns after callback.
    # Since logic updates state before rerun, the next run sees updated map.
    # df is rebuilt with map -> Editor shows new name.
    # Perfect.


    # [v19 New] Display Social Graph
    st.write("---")
    st.subheader("🕸️ 班級社交互動網絡圖 (Social Graph)")
    if st.session_state.id_interactions:
        # Generate graph
        try:
            graph_img = draw_social_graph(st.session_state.id_interactions, 
                                        {m: f"{m}" for m in st.session_state.id_list})
            # [v20.11 Refine] Larger Display Width (1100px)
            # [Fix] Convert BGR to RGB for correct color display in Streamlit
            # Display Graph
            graph_img_rgb = cv2.cvtColor(graph_img, cv2.COLOR_BGR2RGB)
            # [v35 Fix] Removed duplicate st.image call and updated caption per user request
            st.image(graph_img_rgb, use_container_width=True, caption="🔴紅色=社交核心 | 🔵藍色=一般 | 線條粗細=互動頻率")

            # [v34 New] Interaction Details List to clarify connections
            with st.expander("詳細互動清單 (Interaction Details)"):
                if 'id_interactions' in st.session_state and st.session_state.id_interactions:
                    # Sort by count
                    sorted_inters = sorted(st.session_state.id_interactions.items(), key=lambda x: x[1], reverse=True)

                    # [v21.4 Refine] Split list into 2 columns for better readability
                    col_L, col_R = st.columns(2)
                    mid_idx = (len(sorted_inters) + 1) // 2
                    
                    # Prepare ID map for names
                    id_name_map = {}
                    if not df.empty:
                        for _, row in df.iterrows():
                            # Extract ID from display name like "ID_3 (原:3)"
                            r_id_col = row.get('幼兒 ID', "")
                            try:
                                if "ID_" in str(r_id_col):
                                    p = str(r_id_col).split("ID_")[1].split(" ")[0]
                                    int_id = int(p)
                                    id_name_map[int_id] = row['幼兒 ID'] 
                            except:
                                pass
                    for i, (pair, count) in enumerate(sorted_inters):
                        p1, p2 = pair
                        name1 = id_name_map.get(p1, f"ID_{p1}")
                        name2 = id_name_map.get(p2, f"ID_{p2}")
                        
                        # [v35 Fix] Time calculation based on interaction count
                        sec = count / 30.0 # Standard 30fps base
                        
                        target_col = col_L if i < mid_idx else col_R
                        target_col.markdown(f"🔗 **{name1}** ↔ **{name2}**: 互動約 {sec:.1f} 秒 (強度: {count})")

            st.markdown("""
            **📖 如何解讀社交圖譜 (How to Read)**
            *   **🔵 藍色圓圈 (Blue Nodes)**: 一般幼兒 (ID)。
            *   **🔴 紅色圓圈 (Red Nodes)**: 社交核心人物 (互動頻率高於平均)。
            *   **➖ 連線 (Edges)**: 代表兩人間有顯著互動 (如靠近、視線交流)。
            *   **線條粗細**: 線條越粗，代表互動頻率越高。
            """)
        except Exception as e:
            st.error(f"無法繪製社交圖表: {e}")
    else:
        st.info("尚未偵測到顯著的互動事件 (需靠近且持續互動)。")

    if st.button("✨ 點此產生 Excel 報表數據"):
        out = io.BytesIO()
        final_excel_df = edited_df.copy()
        # 下載時將 ID 轉回字串格式，避免 Excel 視為數字
        final_excel_df["幼兒 ID"] = final_excel_df["幼兒 ID"].apply(lambda x: f"ID_{x}" if str(x).isdigit() else str(x))

        final_excel_df.insert(0, "觀察員", observer_name)
        final_excel_df.insert(1, "活動名稱", act_name)
        final_excel_df.insert(2, "觀察日期", act_date)
        final_excel_df.insert(3, "音樂元素", music_element)

        # [v20.1 Update] Advanced Excel Formatting with specific column widths
        # [v20.1 Update] Advanced Excel Formatting with specific column widths
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            final_excel_df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']

            # [v65 Refactor] Dynamic Column Widths based on actual DataFrame columns
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            col_width_map = {
                "觀察員": 12, "活動名稱": 25, "觀察日期": 15, "音樂元素": 10,
                "序號": 5, "幼兒 ID": 25, "AI 服裝特徵": 35, "特徵補強 (圖案/熊/亮片)": 25,
                "AI 觀察判定 (1-5)": 20, "跟隨指令 (同步率%)": 20, "時序延遲 (Lag)": 15,
                "專注度(%)": 12, "動作能量": 12, "動作穩定度": 12, "參與型態": 20,
                "動作檢測 (舉手、側臉)": 25, "AI 總結評語": 60, "教師評分 (1-5)": 15, "教師評語": 30
            }

            wrap_cols = ["AI 服裝特徵", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評語"]

            # Apply widths
            for idx, col_name in enumerate(final_excel_df.columns, 1):
                col_letter = get_column_letter(idx)
                if col_name in col_width_map:
                    worksheet.column_dimensions[col_letter].width = col_width_map[col_name]

            # Apply text wrapping and alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    col_idx = cell.column - 1
                    if col_idx < len(final_excel_df.columns):
                        col_name = final_excel_df.columns[col_idx]
                        if col_name in wrap_cols:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

        st.session_state.excel_ready_data = out.getvalue()
        st.success("🎉 Excel 數據已成功同步！請點擊下方按鈕下載。")

    # [v65 New] Expanded Action Buttons Layout
    col_dl, col_db, col_reset = st.columns([2, 2, 2])

    with col_dl:
        # [v21.4.4 Fix] Double-defensive check for binary data
        ready_data = st.session_state.get('excel_ready_data')
        if ready_data is not None and len(ready_data) > 0:
            st.download_button(
                label="📥 下載 Excel 正式觀察報表",
                data=ready_data,
                file_name=f"HMEAYC_Record_{act_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col_db:
        if st.button("💾 儲存至歷史資料庫 (Save to DB)"):
            # [v37 Fix] Use persistent filename from session_state
            # [v32 Update] Handle return tuple (success, obs_id)
            target_filename = st.session_state.get("current_fn", "unknown_video.mp4")
            success, msg_or_id = save_analysis_to_db(observer_name, act_name, target_filename, edited_df)

            if success:
                st.success(f"✅ 資料已成功儲存！(紀錄 ID: {msg_or_id})")
                st.info("💡 若您再次點擊儲存，系統將會「更新」此筆紀錄，而不會產生重複資料。")
            else:
                st.error(f"❌ 儲存失敗: {msg_or_id}")

    with col_reset:
        # [v65 Restored] Full Reset Button (Clear Cache and Uploads)
        if st.button("🗑️ 清空並建立新分析 (Clear All)"):
            for key in list(st.session_state.keys()):
                # Delete temp file if exists
                if key == 'current_tfile_path':
                    try: os.remove(st.session_state[key])
                    except: pass
                del st.session_state[key]
            st.rerun()
